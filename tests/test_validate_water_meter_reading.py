import unittest
import os
import shutil
from pathlib import Path
import openpyxl
from datetime import datetime, date

from report_builder.validate_water_meter_reading import (
    normalize_flat,
    parse_wateron_filename,
    load_wateron_data,
    find_month_column,
    validate_and_update_water_meter_reading,
    DEFAULT_VARIANCE_THRESHOLD_PCT
)

class TestValidateWaterMeterReading(unittest.TestCase):

    def test_normalize_flat(self):
        self.assertEqual(normalize_flat("F1"), "F1")
        self.assertEqual(normalize_flat("f1"), "F1")
        self.assertEqual(normalize_flat("F 1"), "F1")
        self.assertEqual(normalize_flat("F01"), "F1")
        self.assertEqual(normalize_flat("G09"), "G9")
        self.assertEqual(normalize_flat("C H"), "CH")
        self.assertEqual(normalize_flat("C GYM"), "GYM")
        self.assertEqual(normalize_flat("CBR"), "BSMT")
        self.assertEqual(normalize_flat(""), "")
        self.assertEqual(normalize_flat(None), "")

    def test_parse_wateron_filename(self):
        m, y = parse_wateron_filename("Consumption Report July-2026-ver1-8-10-26.xlsx")
        self.assertEqual((m, y), (7, 2026))

        m, y = parse_wateron_filename("Consumption Report Jul-2026.xlsx")
        self.assertEqual((m, y), (7, 2026))

        m, y = parse_wateron_filename("Consumption Report 07-2026.xlsx")
        self.assertEqual((m, y), (7, 2026))

        m, y = parse_wateron_filename("Consumption Report August 2026.xlsx")
        self.assertEqual((m, y), (8, 2026))

        with self.assertRaises(ValueError):
            parse_wateron_filename("Invalid_Report_Filename.xlsx")

    def test_find_month_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="flat#")
        ws.cell(row=1, column=2, value="flat owner")
        ws.cell(row=1, column=3, value=datetime(2026, 4, 26, 0, 0))
        ws.cell(row=1, column=4, value=datetime(2026, 5, 26, 0, 0))
        ws.cell(row=1, column=5, value=datetime(2026, 6, 26, 0, 0))
        ws.cell(row=1, column=6, value=datetime(2026, 7, 26, 0, 0))
        ws.cell(row=1, column=12, value="average water usage (L)")
        ws.cell(row=1, column=13, value="current")

        col_idx = find_month_column(ws, 7, 2026)
        self.assertEqual(col_idx, 6)

        col_idx_apr = find_month_column(ws, 4, 2026)
        self.assertEqual(col_idx_apr, 3)

        col_idx_notfound = find_month_column(ws, 1, 2025)
        self.assertIsNone(col_idx_notfound)

    def test_end_to_end_validation(self):
        sample_consumption = Path("db/wateron/26-27/Consumption Report July-2026-ver1-8-10-26.xlsx")
        sample_avg = Path("db/wateron/avg-water-usage.xlsx")

        if not (sample_consumption.exists() and sample_avg.exists()):
            self.skipTest("Sample files not present for end-to-end testing")

        test_avg_copy = Path("db/wateron/test_avg_copy.xlsx")
        shutil.copy(sample_avg, test_avg_copy)

        try:
            success = validate_and_update_water_meter_reading(sample_consumption, test_avg_copy)
            self.assertTrue(success)

            wb = openpyxl.load_workbook(test_avg_copy)
            ws = wb.active

            # Verify M1 header
            self.assertEqual(ws.cell(row=1, column=13).value, "Variance %")

            # Verify S13 (Row 46) July reading and high variance highlight
            s13_july = ws.cell(row=46, column=6).value
            self.assertEqual(s13_july, 467.0)

            s13_var_cell = ws.cell(row=46, column=13)
            self.assertEqual(s13_var_cell.value, "=IFERROR((F46-L46)/L46, 0)")
            self.assertEqual(s13_var_cell.number_format, "0.0%")
            self.assertEqual(s13_var_cell.fill.start_color.rgb, "00FFC7CE")

            wb.close()
        finally:
            if test_avg_copy.exists():
                test_avg_copy.unlink()

    def test_custom_threshold_configuration(self):
        sample_consumption = Path("db/wateron/26-27/Consumption Report July-2026-ver1-8-10-26.xlsx")
        sample_avg = Path("db/wateron/avg-water-usage.xlsx")

        if not (sample_consumption.exists() and sample_avg.exists()):
            self.skipTest("Sample files not present for testing")

        test_avg_copy = Path("db/wateron/test_avg_custom_thresh.xlsx")
        shutil.copy(sample_avg, test_avg_copy)

        try:
            # Test with 20% threshold (F14 has +25.2% variance, so it should be highlighted under 20% threshold)
            success = validate_and_update_water_meter_reading(
                sample_consumption,
                test_avg_copy,
                variance_threshold_pct=20.0
            )
            self.assertTrue(success)

            wb = openpyxl.load_workbook(test_avg_copy)
            ws = wb.active

            # F14 is row 15 in avg file
            f14_var_cell = ws.cell(row=15, column=13)
            self.assertEqual(f14_var_cell.fill.start_color.rgb, "00FFC7CE")

            wb.close()
        finally:
            if test_avg_copy.exists():
                test_avg_copy.unlink()

if __name__ == "__main__":
    unittest.main()
