import unittest
import shutil
from pathlib import Path
import openpyxl

from report_builder.update_late_payment_fine import (
    normalize_flat,
    parse_fy_months_from_sheetnames,
    find_sheet,
    update_late_payment_fines,
    ensure_total_expense_formulas,
    get_active_workbook_path
)

class TestUpdateLatePaymentFine(unittest.TestCase):

    def test_normalize_flat(self):
        self.assertEqual(normalize_flat("F1"), "F1")
        self.assertEqual(normalize_flat("f1"), "F1")
        self.assertEqual(normalize_flat("F 1"), "F1")
        self.assertEqual(normalize_flat("F01"), "F1")
        self.assertEqual(normalize_flat("G09"), "G9")
        self.assertEqual(normalize_flat("C H"), "CH")
        self.assertEqual(normalize_flat("C GYM"), "GYM")
        self.assertEqual(normalize_flat("CBR"), "BSMT")

    def test_parse_fy_months_from_sheetnames(self):
        sheets = ["Apr2026-EXPENSE", "Apr2026-COLLECTION", "May2026-EXPENSE", "Jan2027-COLLECTION"]
        months = parse_fy_months_from_sheetnames(sheets)
        self.assertEqual(len(months), 12)
        self.assertEqual(months[0], (4, 2026))   # April 2026
        self.assertEqual(months[1], (5, 2026))   # May 2026
        self.assertEqual(months[11], (3, 2027))  # March 2027

    def test_find_sheet(self):
        wb = openpyxl.Workbook()
        wb.create_sheet("June2026-COLLECTION")
        wb.create_sheet("Jul2026-EXPENSE")

        s1 = find_sheet(wb, ["Jun2026-COLLECTION", "June2026-COLLECTION"])
        self.assertEqual(s1, "June2026-COLLECTION")

        s2 = find_sheet(wb, ["Jul2026-EXPENSE", "July2026-EXPENSE"])
        self.assertEqual(s2, "Jul2026-EXPENSE")

        s3 = find_sheet(wb, ["Aug2026-COLLECTION"])
        self.assertIsNone(s3)

    def test_ensure_total_expense_formulas(self):
        active_wb_path = get_active_workbook_path()
        if not active_wb_path or not active_wb_path.exists():
            self.skipTest("Active accounts workbook not found for testing")

        test_wb_copy = Path("db/accounts/test_formula_unit.xlsx")
        shutil.copy(active_wb_path, test_wb_copy)

        try:
            wb = openpyxl.load_workbook(test_wb_copy, data_only=False)
            updates = ensure_total_expense_formulas(wb)
            wb.save(test_wb_copy)
            wb.close()

            # Verify that formula in May2026-EXPENSE row 6 includes column Q
            wb_check = openpyxl.load_workbook(test_wb_copy, data_only=False)
            ws_may = wb_check["May2026-EXPENSE"]
            headers = [ws_may.cell(2, c).value for c in range(1, ws_may.max_column + 1)]
            total_col = headers.index("TOTAL EXPENSE TO BE PAID") + 1
            formula_r6 = ws_may.cell(6, total_col).value
            self.assertIn("Q6", formula_r6)
            wb_check.close()
        finally:
            if test_wb_copy.exists():
                test_wb_copy.unlink()

    def test_end_to_end_fine_calculation_one_month_lookback(self):
        active_wb_path = get_active_workbook_path()
        if not active_wb_path or not active_wb_path.exists():
            self.skipTest("Active accounts workbook not found for testing")

        test_wb_copy = Path("db/accounts/test_fine_copy.xlsx")
        shutil.copy(active_wb_path, test_wb_copy)

        try:
            success = update_late_payment_fines(test_wb_copy, fine_per_month=1000)
            self.assertTrue(success)

            wb = openpyxl.load_workbook(test_wb_copy, data_only=True)
            
            def get_flat_fine(sheet_name, flat_str):
                ws = wb[sheet_name]
                headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
                fine_col = headers.index("LATE PAYMENT FINE") + 1
                for r in range(3, ws.max_row + 1):
                    if normalize_flat(ws.cell(r, 1).value) == flat_str:
                        val = ws.cell(r, fine_col).value
                        return float(val) if val is not None else 0.0
                return 0.0

            # 1. F2 missed Apr2026-COLLECTION -> May2026-EXPENSE fine = 1000
            self.assertEqual(get_flat_fine("May2026-EXPENSE", "F2"), 1000.0)

            # 2. F2 missed May2026-COLLECTION -> June2026-EXPENSE fine = 1000
            self.assertEqual(get_flat_fine("June2026-EXPENSE", "F2"), 1000.0)

            # 3. F2 paid June2026-COLLECTION -> July2026-EXPENSE fine = 0
            self.assertEqual(get_flat_fine("July2026-EXPENSE", "F2"), 0.0)

            wb.close()
        finally:
            if test_wb_copy.exists():
                test_wb_copy.unlink()

    def test_future_months_skipped_or_zero(self):
        active_wb_path = get_active_workbook_path()
        if not active_wb_path or not active_wb_path.exists():
            self.skipTest("Active accounts workbook not found for testing")

        test_wb_copy = Path("db/accounts/test_future_fine_copy.xlsx")
        shutil.copy(active_wb_path, test_wb_copy)

        try:
            success = update_late_payment_fines(test_wb_copy, fine_per_month=1000, as_of_date="2026-08")
            self.assertTrue(success)

            wb = openpyxl.load_workbook(test_wb_copy, data_only=True)
            
            def get_flat_fine(sheet_name, flat_str):
                ws = wb[sheet_name]
                headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
                fine_col = headers.index("LATE PAYMENT FINE") + 1
                for r in range(3, ws.max_row + 1):
                    if normalize_flat(ws.cell(r, 1).value) == flat_str:
                        val = ws.cell(r, fine_col).value
                        return float(val) if val is not None else 0.0
                return 0.0

            # Current month (Aug2026-EXPENSE) and future months MUST have 0 fine for all flats
            self.assertEqual(get_flat_fine("Aug2026-EXPENSE", "F2"), 0.0)
            self.assertEqual(get_flat_fine("Sep2026-EXPENSE", "F1"), 0.0)
            self.assertEqual(get_flat_fine("Sep2026-EXPENSE", "F2"), 0.0)
            self.assertEqual(get_flat_fine("Jan2027-EXPENSE", "F2"), 0.0)

            wb.close()
        finally:
            if test_wb_copy.exists():
                test_wb_copy.unlink()

if __name__ == "__main__":
    unittest.main()

