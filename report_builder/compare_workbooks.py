#!/usr/bin/env python3
"""
Compare formulas between a baseline and candidate workbook.

Reports all formula differences: missing formulas, structural mismatches,
and formulas replaced by static values.

Usage:
    python compare_workbooks.py <baseline.xlsx> <candidate.xlsx>

Example:
    python compare_workbooks.py \
        ../db/accounts/2025-2026-INCOME-EXPENDITURE-ACCOUNT-4.26.26-gold.xlsx \
        ../db/accounts/2026-2027-INCOME-EXPENDITURE-ACCOUNT-5.03.26-gold.xlsx
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


def detect_fy_start(wb: openpyxl.Workbook) -> int | None:
    for name in wb.sheetnames:
        if name.startswith("Apr") and name.endswith("-EXPENSE"):
            mid = name.replace("Apr", "").replace("-EXPENSE", "")
            if mid.isdigit():
                return int(mid)
    return None


def build_sheet_map(old_fy_start: int) -> dict[str, str]:
    new_apr = old_fy_start + 1
    old_jan = old_fy_start + 1
    new_jan = old_fy_start + 2

    renames: dict[str, str] = {}
    months_alt = ["Apr", "May", "Aug", "Sep", "Oct", "Nov", "Dec"]
    suffixes = ["-EXPENSE", "-COLLECTION"]

    for m in months_alt:
        for sfx in suffixes:
            renames[f"{m}{old_fy_start}{sfx}"] = f"{m}{new_apr}{sfx}"
    for sfx in suffixes:
        renames[f"June{old_fy_start}{sfx}"] = f"June{new_apr}{sfx}"
        renames[f"July{old_fy_start}{sfx}"] = f"July{new_apr}{sfx}"
    for extra in ["-EXPENSE-Flatwise", "-EXPENSE-Flatwise-Trans"]:
        renames[f"July{old_fy_start}{extra}"] = f"July{new_apr}{extra}"
    for m in ["Jan", "Feb", "Mar"]:
        for sfx in suffixes:
            renames[f"{m}{old_jan}{sfx}"] = f"{m}{new_jan}{sfx}"
    renames[f"INCOME-EXPENSE-CYCLES-{old_fy_start}-{old_fy_start + 1 - 2000}"] = (
        f"INCOME-EXPENSE-CYCLES-{new_apr}-{new_apr + 1 - 2000}"
    )
    return renames


def normalize_formula(formula: str, fy_start: int) -> str:
    if not isinstance(formula, str):
        return str(formula)

    # Skip DUMMYFUNCTION formulas — these are Google Sheets cached values
    # and contain arbitrary numeric literals that shouldn't be year-normalized
    if "DUMMYFUNCTION" in formula:
        return formula

    s = formula
    apr_year = fy_start
    jan_year = fy_start + 1

    months_jan = ["January", "February", "March", "Jan", "Feb", "Mar"]
    months_apr = [
        "April", "August", "September", "October", "November", "December",
        "June", "July", "Apr", "May", "Aug", "Sep", "Oct", "Nov", "Dec",
    ]

    for m in months_jan:
        s = s.replace(f"{m}{jan_year}", f"{m}YYYY2")
        s = s.replace(f"{m} {jan_year}", f"{m} YYYY2")
    for m in months_apr:
        s = s.replace(f"{m}{apr_year}", f"{m}YYYY1")
        s = s.replace(f"{m} {apr_year}", f"{m} YYYY1")

    s = s.replace(str(jan_year), "YYYY2")
    s = s.replace(str(apr_year), "YYYY1")
    return s


def compare(baseline_path: str, candidate_path: str) -> int:
    print(f"Baseline:  {baseline_path}")
    print(f"Candidate: {candidate_path}")
    print()

    wb_base = openpyxl.load_workbook(baseline_path)
    wb_cand = openpyxl.load_workbook(candidate_path)

    base_fy = detect_fy_start(wb_base)
    cand_fy = detect_fy_start(wb_cand)

    if base_fy is None or cand_fy is None:
        print("ERROR: Could not detect FY start from sheet names.")
        return 1

    print(f"Baseline FY:  {base_fy}-{(base_fy + 1) % 100:02d}")
    print(f"Candidate FY: {cand_fy}-{(cand_fy + 1) % 100:02d}")
    print()

    sheet_map = build_sheet_map(base_fy) if base_fy != cand_fy else {}
    for sn in wb_base.sheetnames:
        if sn not in sheet_map:
            sheet_map[sn] = sn

    missing_sheets: list[str] = []
    missing_formulas: dict[str, list[tuple[str, str]]] = defaultdict(list)
    mismatches: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    replaced_by_value: dict[str, list[tuple[str, str, object]]] = defaultdict(list)
    total_checked = 0
    total_ok = 0

    for base_sheet, cand_sheet in sheet_map.items():
        if base_sheet not in wb_base.sheetnames:
            continue
        if cand_sheet not in wb_cand.sheetnames:
            missing_sheets.append(f"{cand_sheet} (expected from {base_sheet})")
            continue

        ws_base = wb_base[base_sheet]
        ws_cand = wb_cand[cand_sheet]

        for row in ws_base.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                total_checked += 1
                coord = cell.coordinate
                old_formula = cell.value

                new_cell = ws_cand[coord]
                new_value = new_cell.value

                if new_value is None or new_value == 0 or new_value == "":
                    missing_formulas[cand_sheet].append((coord, old_formula))
                elif not (isinstance(new_value, str) and new_value.startswith("=")):
                    replaced_by_value[cand_sheet].append((coord, old_formula, new_value))
                else:
                    norm_old = normalize_formula(old_formula, base_fy)
                    norm_new = normalize_formula(new_value, cand_fy)
                    if norm_old != norm_new:
                        mismatches[cand_sheet].append((coord, old_formula, new_value))
                    else:
                        total_ok += 1

    total_missing = sum(len(v) for v in missing_formulas.values())
    total_mismatch = sum(len(v) for v in mismatches.values())
    total_replaced = sum(len(v) for v in replaced_by_value.values())
    total_issues = total_missing + total_mismatch + total_replaced + len(missing_sheets)

    print("=" * 80)
    print("FORMULA COMPARISON REPORT")
    print("=" * 80)
    print(f"  Total formulas checked:           {total_checked}")
    print(f"  Correctly shifted / matching:      {total_ok}")
    print(f"  Missing formulas (now blank/zero): {total_missing}")
    print(f"  Structural mismatches:             {total_mismatch}")
    print(f"  Replaced by static value:          {total_replaced}")
    print(f"  Missing sheets:                    {len(missing_sheets)}")
    print()

    if total_issues == 0:
        print("ALL FORMULAS MATCH. No issues found.")
        return 0

    if missing_sheets:
        print("-" * 80)
        print("MISSING SHEETS")
        print("-" * 80)
        for s in missing_sheets:
            print(f"  {s}")
        print()

    if missing_formulas:
        print("-" * 80)
        print("MISSING FORMULAS (formula in baseline, blank/zero in candidate)")
        print("-" * 80)
        for sheet, items in sorted(missing_formulas.items()):
            print(f"\n  [{sheet}] — {len(items)} missing")
            for coord, old_f in items[:10]:
                print(f"    {coord}: {old_f[:90]}")
            if len(items) > 10:
                print(f"    ... and {len(items) - 10} more")
        print()

    if mismatches:
        print("-" * 80)
        print("FORMULA MISMATCHES (structure differs after year-normalization)")
        print("-" * 80)
        for sheet, items in sorted(mismatches.items()):
            print(f"\n  [{sheet}] — {len(items)} mismatched")
            for coord, old_f, new_f in items[:10]:
                print(f"    {coord}:")
                print(f"      Baseline:  {old_f[:100]}")
                print(f"      Candidate: {new_f[:100]}")
            if len(items) > 10:
                print(f"    ... and {len(items) - 10} more")
        print()

    if replaced_by_value:
        print("-" * 80)
        print("FORMULAS REPLACED BY STATIC VALUES")
        print("-" * 80)
        for sheet, items in sorted(replaced_by_value.items()):
            print(f"\n  [{sheet}] — {len(items)} replaced")
            for coord, old_f, new_v in items[:10]:
                print(f"    {coord}: was '{old_f[:60]}' → now '{str(new_v)[:40]}'")
            if len(items) > 10:
                print(f"    ... and {len(items) - 10} more")
        print()

    print("=" * 80)
    print(f"TOTAL ISSUES: {total_issues}")
    print("=" * 80)
    return 1 if total_issues > 0 else 0


def main():
    parser = argparse.ArgumentParser(
        description="Compare formulas between baseline and candidate workbooks"
    )
    parser.add_argument("baseline", help="Path to the baseline (reference) .xlsx workbook")
    parser.add_argument("candidate", help="Path to the candidate .xlsx workbook to check")
    args = parser.parse_args()

    if not Path(args.baseline).exists():
        print(f"ERROR: Baseline file not found: {args.baseline}")
        sys.exit(1)
    if not Path(args.candidate).exists():
        print(f"ERROR: Candidate file not found: {args.candidate}")
        sys.exit(1)

    sys.exit(compare(args.baseline, args.candidate))


if __name__ == "__main__":
    main()
