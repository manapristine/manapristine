"""
Main Society Financial Report Builder Engine

Overview:
---------
This script is the core report processing engine of the Mana Pristine financial management system.
It reads society financial workbooks configured in 'db/workbooks.json', processes flat-wise
income and expenditure records across monthly collection and expense sub-sheets, and generates
static JSON report datasets in the 'docs/' directory for web dashboard consumption.

Pre-conditions:
---------------
1. Configuration File:
   'db/workbooks.json' MUST exist and specify financial year mappings to workbook relative paths
   (e.g., "2026-27": {"workbook": "db/accounts/2026-2027-INCOME-EXPENDITURE-ACCOUNT-8.10.26-gold.xlsx"}).

2. Mapping Files:
   - 'db/members.csv' MUST exist with headers 'flat', 'name', and 'email' (owner mappings).
   - 'db/occupants.csv' MUST exist with headers 'flat' and 'name' (resident/occupant mappings).

3. Workbook Sub-sheet Structure:
   - Each financial workbook MUST contain an 'INCOME-EXPENSE-CYCLES' sheet or monthly collection
     sheets ('<month>-COLLECTION') and expense sheets ('<month>-EXPENSE').
   - Expense sheets MUST contain columns 'LATE PAYMENT FINE' (for flat-wise late fees) and
     'TOTAL EXPENSE TO BE PAID' (for total monthly expenses).
   - An optional 'ANNUAL-EXPENSE-DETAILS' sheet can be present for line-item expense category breakdowns.

4. Output Directory Access:
   The 'docs/' directory MUST be writable to generate 'report-data-<FY>.json' and 'report-manifest.json'.

Workflow & Output:
------------------
1. Parses flat owner and occupant mappings from 'db/members.csv' and 'db/occupants.csv'.
2. Reads financial year workbooks from 'db/workbooks.json'.
3. Extracts monthly collection totals, expense totals, and late payment fines directly from the
   'LATE PAYMENT FINE' column of each respective '<month>-EXPENSE' sheet.
4. Hashes the community portal password (if present in workbooks.json) and injects SHA-256 hash into 'docs/index.html'.
5. Writes processed flat-wise financial statements to 'docs/report-data-<FY>.json'.
6. Writes 'docs/report-manifest.json' listing all available report datasets.

CLI Usage:
----------
1. Standard Execution (processes all configured workbooks):
   python report_builder/report_builder.py

2. Help Menu:
   python report_builder/report_builder.py --help

Exit Codes:
-----------
  0: Success (report datasets and manifest successfully generated)
  1: Failure (missing required files or invalid workbook configuration)
"""

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook



PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SHEET = "INCOME-EXPENSE-CYCLES"
WORKBOOKS_JSON = PROJECT_ROOT / "db" / "workbooks.json"
MEMBERS_CSV = PROJECT_ROOT / "db" / "members.csv"
OCCUPANTS_CSV = PROJECT_ROOT / "db" / "occupants.csv"
OUTPUT_DIR = PROJECT_ROOT / "docs"


@dataclass(frozen=True)
class MonthlyBlock:
    month_label: str
    collection_idx: int
    expense_idx: int
    late_fee_idx: int
    net_idx: int


@dataclass(frozen=True)
class SheetLayout:
    carry_over_idx: int | None
    total_collection_idx: int | None
    total_expense_idx: int | None
    total_late_fee_idx: int | None
    total_net_idx: int | None
    total_dues_idx: int | None


def update_portal_password(password: str):
    """Hash the portal password and inject it into docs/index.html."""
    if not password:
        return
    portal_hash = hashlib.sha256(password.encode()).hexdigest()
    index_path = OUTPUT_DIR / "index.html"
    if not index_path.exists():
        print(f"Warning: index.html not found at {index_path}")
        return
    
    content = index_path.read_text(encoding="utf-8")
    new_content = re.sub(
        r'const PORTAL_HASH = "[^"]*";', 
        f'const PORTAL_HASH = "{portal_hash}";', 
        content
    )
    if content != new_content:
        index_path.write_text(new_content, encoding="utf-8")
        print(f"Portal password updated in {index_path}")
    else:
        print("Portal password hash is already up to date.")


def normalize_flat(flat: str) -> str:
    return (flat or "").strip().upper()


def safe_number(value: Any) -> float | int:
    if value is None or value == "":
        return 0
    return value


def month_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%b %Y")
    return str(value).strip()


def derive_financial_year(monthly_blocks: list[MonthlyBlock]) -> str:
    """Derive financial year string like '2025-26' from monthly block headers."""
    if not monthly_blocks:
        return "Unknown"
    first = monthly_blocks[0].month_label   # e.g., "Apr 2025"
    last = monthly_blocks[-1].month_label   # e.g., "Mar 2026"
    start_year = first.split()[-1]
    end_year = last.split()[-1]
    if start_year == end_year:
        return start_year
    return f"{start_year}-{end_year[-2:]}"


def build_expense_sheet_map(
    monthly_blocks: list[MonthlyBlock], available_sheets: list[str]
) -> dict[str, str]:
    """Build month_label -> sheet_name map by matching available workbook sheets."""
    sheet_set = set(available_sheets)
    full_month = {
        "Jan": "January", "Feb": "February", "Mar": "March",
        "Apr": "April", "May": "May", "Jun": "June",
        "Jul": "July", "Aug": "August", "Sep": "September",
        "Oct": "October", "Nov": "November", "Dec": "December",
    }
    result: dict[str, str] = {}
    for block in monthly_blocks:
        parts = block.month_label.split()
        if len(parts) != 2:
            continue
        abbrev, year = parts
        candidates = [f"{abbrev}{year}-EXPENSE"]
        full = full_month.get(abbrev, "")
        if full and full != abbrev:
            candidates.append(f"{full}{year}-EXPENSE")
        for candidate in candidates:
            if candidate in sheet_set:
                result[block.month_label] = candidate
                break
    return result


def write_report_dataset(
    reports: list[dict[str, Any]],
    output_json: Path,
    financial_year: str,
    annual_expense_details: list[dict[str, Any]] | None = None,
    cutoff_date: str | None = None,
) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "financial_year": financial_year,
        "cutoff_date": cutoff_date,
        "sheet": DEFAULT_SHEET,
        "report_count": len(reports),
        "reports": reports,
    }
    if annual_expense_details is not None:
        payload["annual_expense_details"] = annual_expense_details
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def load_flat_requests(members_csv: Path) -> list[dict[str, str]]:
    """Load the list of flats and owner info directly from members.csv."""
    flat_requests: list[dict[str, str]] = []
    with members_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            flat = normalize_flat(row.get("flat", ""))
            if not flat:
                continue
            flat_requests.append(
                {
                    "flat": flat,
                    "owner_name": (row.get("name") or "").strip(),
                    "email": (row.get("email") or "").strip(),
                }
            )
    return flat_requests


def load_occupants(occupants_csv: Path) -> dict[str, str]:
    """Return a mapping of normalized flat -> occupant name from occupants.csv."""
    lookup: dict[str, str] = {}
    with occupants_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            flat = normalize_flat(row.get("flat", ""))
            if flat:
                lookup[flat] = (row.get("name") or "").strip()
    return lookup


def extract_monthly_blocks(header_row_1: tuple[Any, ...], header_row_2: tuple[Any, ...]) -> list[MonthlyBlock]:
    blocks: list[MonthlyBlock] = []
    for idx, cell_value in enumerate(header_row_1):
        if not isinstance(cell_value, datetime):
            continue
        if idx + 3 >= len(header_row_2):
            continue
        labels = [header_row_2[idx + offset] for offset in range(4)]
        if labels != ["COLLECTION", "EXPENSE", "LATE PAYMENT FEE", "COLLECTION - EXPENSE"]:
            continue
        blocks.append(
            MonthlyBlock(
                month_label=month_label(cell_value),
                collection_idx=idx,
                expense_idx=idx + 1,
                late_fee_idx=idx + 2,
                net_idx=idx + 3,
            )
        )
    return blocks


def extract_sheet_layout(header_row_1: tuple[Any, ...], header_row_2: tuple[Any, ...]) -> SheetLayout:
    def find_idx(patterns: list[str]) -> int | None:
        # Search in Row 1
        for idx, val in enumerate(header_row_1):
            s = str(val or "").strip().replace("\n", " ")
            if any(p in s for p in patterns):
                return idx
        # Search in Row 2
        for idx, val in enumerate(header_row_2):
            s = str(val or "").strip().replace("\n", " ")
            if any(p in s for p in patterns):
                return idx
        return None

    return SheetLayout(
        carry_over_idx=find_idx(["Balance from last FY year"]),
        total_collection_idx=find_idx(["TOTAL COLLECTION"]),
        total_expense_idx=find_idx(["TOTAL EXPENSE"]),
        total_late_fee_idx=find_idx(["TOTAL LATE FEE"]),
        total_net_idx=find_idx(["TOTAL COLLECTION - EXPENSE"]),
        total_dues_idx=find_idx(["Total Dues"]),
    )


def load_collection_totals(workbook_path: Path, sheet_name: str) -> dict[str, float]:
    """Read a COLLECTION sheet and compute totals per flat (sum of amount columns)."""
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}
        ws = workbook[sheet_name]
        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            flat = normalize_flat(row[0] if row else "")
            if not flat:
                continue
            total = 0.0
            for col_idx in [2, 4, 6, 8]:
                val = row[col_idx] if len(row) > col_idx else None
                if isinstance(val, (int, float)):
                    total += val
            result[flat] = total
        return result
    finally:
        workbook.close()


def load_expense_totals(workbook_path: Path, sheet_name: str) -> dict[str, float]:
    """Read total expense per flat directly from cached cell values ('TOTAL EXPENSE TO BE PAID')."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}
        ws = workbook[sheet_name]
        headers = [str(cell.value).upper() if cell.value else "" for cell in ws[2]]
        total_col_idx = -1
        for idx, h in enumerate(headers):
            if "TOTAL EXPENSE TO BE PAID" in h:
                total_col_idx = idx
                break
        if total_col_idx == -1:
            total_col_idx = 17

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            flat = normalize_flat(row[0] if row else "")
            if not flat or flat in ("CH", "GYM", "BSMT", "MPFOWA", "TOTAL"):
                continue
            total_expense = row[total_col_idx] if len(row) > total_col_idx else None
            if total_expense is None or not isinstance(total_expense, (int, float)):
                raise ValueError(
                    f"Null or missing cached formula value for 'TOTAL EXPENSE TO BE PAID' in workbook '{workbook_path.name}', "
                    f"sheet '{sheet_name}', flat '{flat}'. Please open the workbook in Excel, save it, close it, and re-run."
                )
            result[flat] = float(total_expense)
        return result
    finally:
        workbook.close()


def load_late_fee_totals(workbook_path: Path, sheet_name: str) -> dict[str, float]:
    """Read late payment fine per flat directly from the 'LATE PAYMENT FINE' column in an EXPENSE sheet."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}
        ws = workbook[sheet_name]
        headers = [str(cell.value).upper() if cell.value else "" for cell in ws[2]]
        fine_col_idx = -1
        for i, h in enumerate(headers):
            if any(kw in h for kw in ["LATE PAYMENT FINE", "LATE PAYMENT FEE", "LATE FINE"]):
                fine_col_idx = i
                break
        if fine_col_idx == -1:
            return {}

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            flat = normalize_flat(row[0] if row else "")
            if not flat or flat in ("CH", "GYM", "BSMT", "MPFOWA", "TOTAL"):
                continue
            fine_val = row[fine_col_idx] if len(row) > fine_col_idx else 0
            result[flat] = float(fine_val) if isinstance(fine_val, (int, float)) else 0.0
        return result
    finally:
        workbook.close()


def get_aed_column_map(ws) -> tuple[dict[str, int], list[tuple[int, str]]]:
    """Identify summary columns and line item columns from ANNUAL-EXPENSE-DETAILS headers."""
    header_row_1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_row_2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    col_names: list[tuple[int, str]] = []
    summary_indices: dict[str, int] = {}

    summary_markers = {
        "gross_expense": ["gross expense"],
        "gross_variable_expense": ["gross variable", "gross var"],
        "gross_fixed_expense": ["gross fixed"],
        "water_meter_rent": ["water meter rent", "meter rent"],
        "total_expense": ["total expense"],
    }

    max_cols = max(len(header_row_1), len(header_row_2))
    for ci in range(max_cols):
        h1 = str(header_row_1[ci]).strip().lower() if ci < len(header_row_1) and header_row_1[ci] else ""
        h2 = str(header_row_2[ci]).strip() if ci < len(header_row_2) and header_row_2[ci] else ""

        matched_summary = False
        for key, markers in summary_markers.items():
            if any(m in h1 for m in markers):
                summary_indices[key] = ci
                matched_summary = True
                break

        if not matched_summary and h2:
            col_names.append((ci, h2))

    return summary_indices, col_names


def load_sheet_rows(
    workbook_path: Path, sheet_name: str
) -> tuple[list[MonthlyBlock], SheetLayout, dict[str, tuple[Any, ...]], list[str]]:
    """Load cell values from the specified summary sheet, layout indices, and monthly blocks."""
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        available_sheets = list(workbook.sheetnames)
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in {workbook_path}")
            return [], SheetLayout(None, None, None, None, None, None), {}, available_sheets

        ws = workbook[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 2:
            return [], SheetLayout(None, None, None, None, None, None), {}, available_sheets

        header_row_1 = all_rows[0]
        header_row_2 = all_rows[1]

        monthly_blocks: list[MonthlyBlock] = []
        current_month: str | None = None
        coll_idx: int | None = None
        exp_idx: int | None = None
        late_fee_idx: int | None = None
        net_idx: int | None = None

        for ci, val in enumerate(header_row_1):
            if val is not None and str(val).strip() != "":
                if current_month is not None and coll_idx is not None and exp_idx is not None:
                    monthly_blocks.append(
                        MonthlyBlock(
                            month_label=current_month,
                            collection_idx=coll_idx,
                            expense_idx=exp_idx,
                            late_fee_idx=late_fee_idx if late_fee_idx is not None else -1,
                            net_idx=net_idx if net_idx is not None else exp_idx + 1,
                        )
                    )
                current_month = month_label(val)
                coll_idx = None
                exp_idx = None
                late_fee_idx = None
                net_idx = None

            if current_month is not None and ci < len(header_row_2):
                h2 = str(header_row_2[ci]).strip().upper() if header_row_2[ci] is not None else ""
                if "COLLECTION" in h2 and "MINUS" not in h2:
                    coll_idx = ci
                elif "EXPENSE" in h2 and "MINUS" not in h2:
                    exp_idx = ci
                elif "LATE" in h2:
                    late_fee_idx = ci
                elif "MINUS" in h2 or "NET" in h2 or "COLLECTION - EXPENSE" in h2:
                    net_idx = ci

        if current_month is not None and coll_idx is not None and exp_idx is not None:
            monthly_blocks.append(
                MonthlyBlock(
                    month_label=current_month,
                    collection_idx=coll_idx,
                    expense_idx=exp_idx,
                    late_fee_idx=late_fee_idx if late_fee_idx is not None else -1,
                    net_idx=net_idx if net_idx is not None else exp_idx + 1,
                )
            )

        labels = [str(header_row_2[ci]).strip().upper() for ci in range(2, 6) if ci < len(header_row_2)]
        if labels != ["COLLECTION", "EXPENSE", "LATE PAYMENT FEE", "COLLECTION - EXPENSE"]:
            pass

        def find_idx(target_headers: list[str]) -> int | None:
            for ci, val in enumerate(header_row_1):
                if val is not None:
                    v_str = str(val).strip().upper()
                    if any(t in v_str for t in target_headers):
                        return ci
            for ci, val in enumerate(header_row_2):
                if val is not None:
                    v_str = str(val).strip().upper()
                    if any(t in v_str for t in target_headers):
                        return ci
            return None

        sheet_layout = SheetLayout(
            carry_over_idx=find_idx(["CLOSING DUES", "CARRY OVER", "BALANCE FROM LAST FY"]),
            total_collection_idx=find_idx(["TOTAL COLLECTION"]),
            total_expense_idx=find_idx(["TOTAL EXPENSE"]),
            total_late_fee_idx=find_idx(["TOTAL LATE FEE"]),
            total_net_idx=find_idx(["TOTAL NET", "TOTAL COLLECTION - EXPENSE"]),
            total_dues_idx=find_idx(["TOTAL DUES", "CLOSING BALANCE"]),
        )

        row_lookup: dict[str, tuple[Any, ...]] = {}
        for row in all_rows[2:]:
            if not row or row[0] is None:
                continue
            flat = normalize_flat(str(row[0]))
            if not flat or flat in ("TOTAL", "GRAND TOTAL", "FLAT #"):
                continue
            row_lookup[flat] = row

        return monthly_blocks, sheet_layout, row_lookup, available_sheets
    finally:
        workbook.close()


def has_cached_values(row_lookup: dict[str, tuple[Any, ...]], monthly_blocks: list[MonthlyBlock]) -> bool:
    if not row_lookup or not monthly_blocks:
        return False
    sample_row = next(iter(row_lookup.values()))
    first_block = monthly_blocks[0]
    coll_val = sample_row[first_block.collection_idx] if len(sample_row) > first_block.collection_idx else None
    exp_val = sample_row[first_block.expense_idx] if len(sample_row) > first_block.expense_idx else None
    return isinstance(coll_val, (int, float)) and isinstance(exp_val, (int, float)) and exp_val > 0


def fill_from_source_sheets(
    workbook_path: Path,
    monthly_blocks: list[MonthlyBlock],
    sheet_layout: SheetLayout,
    row_lookup: dict[str, tuple[Any, ...]],
    available_sheets: list[str],
) -> dict[str, tuple[Any, ...]]:
    """When INCOME-EXPENSE-CYCLES has no cached values, read source sheets directly."""
    expense_sheet_map = build_expense_sheet_map(monthly_blocks, available_sheets)
    collection_sheet_map: dict[str, str] = {}
    for block in monthly_blocks:
        parts = block.month_label.split()
        if len(parts) != 2:
            continue
        abbrev, year = parts
        full_month = {
            "Jan": "January", "Feb": "February", "Mar": "March",
            "Apr": "April", "May": "May", "Jun": "June",
            "Jul": "July", "Aug": "August", "Sep": "September",
            "Oct": "October", "Nov": "November", "Dec": "December",
        }
        candidates = [f"{abbrev}{year}-COLLECTION"]
        full = full_month.get(abbrev, "")
        if full and full != abbrev:
            candidates.append(f"{full}{year}-COLLECTION")
        for candidate in candidates:
            if candidate in available_sheets:
                collection_sheet_map[block.month_label] = candidate
                break

    monthly_collections: dict[str, dict[str, float]] = {}
    for month_label, coll_sheet in collection_sheet_map.items():
        monthly_collections[month_label] = load_collection_totals(workbook_path, coll_sheet)

    monthly_expenses: dict[str, dict[str, float]] = {}
    monthly_late_fees: dict[str, dict[str, float]] = {}
    for month_label, exp_sheet in expense_sheet_map.items():
        monthly_expenses[month_label] = load_expense_totals(workbook_path, exp_sheet)
        monthly_late_fees[month_label] = load_late_fee_totals(workbook_path, exp_sheet)

    new_lookup: dict[str, tuple[Any, ...]] = {}
    for flat, original_row in row_lookup.items():
        row_list = list(original_row)
        while len(row_list) <= max(
            (sheet_layout.total_dues_idx or 0),
            (monthly_blocks[-1].net_idx if monthly_blocks else 0),
        ):
            row_list.append(None)

        for block in monthly_blocks:
            coll = monthly_collections.get(block.month_label, {}).get(flat, 0)
            exp = monthly_expenses.get(block.month_label, {}).get(flat, 0)
            late_fee = monthly_late_fees.get(block.month_label, {}).get(flat, 0)
            row_list[block.collection_idx] = coll
            row_list[block.expense_idx] = exp
            row_list[block.late_fee_idx] = late_fee
            row_list[block.net_idx] = coll - exp

        total_coll = sum(safe_number(row_list[b.collection_idx]) for b in monthly_blocks)
        total_exp = sum(safe_number(row_list[b.expense_idx]) for b in monthly_blocks)
        total_late = sum(safe_number(row_list[b.late_fee_idx]) for b in monthly_blocks)
        total_net = total_coll - (total_exp + total_late)
        carry_over = safe_number(row_list[sheet_layout.carry_over_idx] if sheet_layout.carry_over_idx is not None and len(row_list) > sheet_layout.carry_over_idx else 0)

        if sheet_layout.total_collection_idx is not None:
            row_list[sheet_layout.total_collection_idx] = total_coll
        if sheet_layout.total_expense_idx is not None:
            row_list[sheet_layout.total_expense_idx] = total_exp
        if sheet_layout.total_late_fee_idx is not None:
            row_list[sheet_layout.total_late_fee_idx] = total_late
        if sheet_layout.total_net_idx is not None:
            row_list[sheet_layout.total_net_idx] = total_net
        if sheet_layout.total_dues_idx is not None:
            row_list[sheet_layout.total_dues_idx] = total_net + carry_over

        new_lookup[flat] = tuple(row_list)

    return new_lookup


def load_monthly_accounting_fees(workbook_path: Path) -> dict[str, float]:
    """Read month_label -> per_flat_accounting_fee (accounting / 64.0) from ANNUAL-EXPENSE-DETAILS."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    res: dict[str, float] = {}
    try:
        if "ANNUAL-EXPENSE-DETAILS" not in workbook.sheetnames:
            return res
        ws = workbook["ANNUAL-EXPENSE-DETAILS"]
        row2 = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[2]]
        acct_idx = -1
        for i, h in enumerate(row2):
            if "accounting" in h:
                acct_idx = i
                break
        if acct_idx == -1:
            return res

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            month_raw = str(row[0]).strip()
            parts = month_raw.split()
            if len(parts) != 2:
                continue
            m_label = f"{parts[0].capitalize()[:3]} {parts[1]}"
            
            acct_val = row[acct_idx] if len(row) > acct_idx else 0
            if isinstance(acct_val, (int, float)) and acct_val > 0:
                res[m_label] = float(acct_val) / 64.0
            else:
                res[m_label] = 0.0
        return res
    finally:
        workbook.close()


def load_expense_details(workbook_path: Path, expense_sheet_map: dict[str, str]) -> dict[str, dict[str, dict[str, float]]]:
    """Read per-flat, per-month expense breakdown directly from cached cell values in the spreadsheet."""
    monthly_acct_fees = load_monthly_accounting_fees(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    result: dict[str, dict[str, dict[str, float]]] = {}
    non_flat_ids = {"CH", "GYM", "BSMT", "MPFOWA", "TOTAL"}
    try:
        for month_label, sheet_name in expense_sheet_map.items():
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            accounts_fee_per_flat = monthly_acct_fees.get(month_label, 0.0)

            headers = [str(cell.value).upper() if cell.value else "" for cell in ws[2]]
            mem_fee_idx = -1
            late_fee_idx = -1
            reason_idx = -1
            total_exp_idx = -1
            for i, h in enumerate(headers):
                if "ANNUAL MEMBERSHIP" in h or "ANNUAL MEM" in h:
                    mem_fee_idx = i
                elif any(kw in h for kw in ["LATE PAYMENT FINE REASON", "LATE FINE REASON"]):
                    reason_idx = i
                elif any(kw in h for kw in ["LATE PAYMENT FINE", "LATE PAYMENT FEE", "LATE FINE"]):
                    late_fee_idx = i
                elif "TOTAL EXPENSE TO BE PAID" in h:
                    total_exp_idx = i

            if total_exp_idx == -1:
                total_exp_idx = 17

            for row in ws.iter_rows(min_row=3, values_only=True):
                flat = normalize_flat(row[0] if row else "")
                if not flat or flat in non_flat_ids:
                    continue

                def _req_num(idx: int, field_name: str) -> float:
                    if idx < 0:
                        return 0.0
                    v = row[idx] if len(row) > idx else None
                    if v is None or not isinstance(v, (int, float)):
                        raise ValueError(
                            f"Missing or null cached value for field '{field_name}' in workbook '{workbook_path.name}', "
                            f"sheet '{sheet_name}', flat '{flat}'. Please open the workbook in Excel, save it, close it, and re-run."
                        )
                    return float(v)

                def _opt_num(idx: int) -> float:
                    if idx < 0:
                        return 0.0
                    v = row[idx] if len(row) > idx else 0
                    return float(v) if isinstance(v, (int, float)) else 0.0

                water_used = _opt_num(2)
                water_expense = _req_num(6, "WATER EXPENSE SHARING BY %")
                num_meters = _opt_num(7)
                meter_rent = _opt_num(8)
                total_water_expense = _opt_num(9)
                fixed_expense = _req_num(10, "FIXED EXPENSE SHARE")
                parking_fee = _opt_num(11)
                club_house_fee = _opt_num(12)
                shifting_fee = _opt_num(13)
                gym_usage_fee = _opt_num(14)
                accounts_fee = accounts_fee_per_flat
                annual_mem_fee = _opt_num(mem_fee_idx)
                late_fee = _opt_num(late_fee_idx)
                total_expense = _req_num(total_exp_idx, "TOTAL EXPENSE TO BE PAID")

                reason_val = ""
                if reason_idx >= 0 and len(row) > reason_idx:
                    r_raw = row[reason_idx]
                    reason_val = str(r_raw).strip() if r_raw is not None else ""

                result.setdefault(flat, {})[month_label] = {
                    "water_used_litres": water_used,
                    "common_area_water_litres": _opt_num(3),
                    "total_fresh_water_consumed_litres": _opt_num(4),
                    "water_expense": water_expense,
                    "num_meters": num_meters,
                    "meter_rent": meter_rent,
                    "total_water_expense": total_water_expense,
                    "fixed_expense": fixed_expense,
                    "parking_fee": parking_fee,
                    "club_house_fee": club_house_fee,
                    "shifting_fee": shifting_fee,
                    "gym_usage_fee": gym_usage_fee,
                    "accounts_fee": accounts_fee,
                    "annual_mem_fee": annual_mem_fee,
                    "late_fee": late_fee,
                    "late_fee_reason": reason_val,
                    "total_expense": total_expense,
                }

        return result
    finally:
        workbook.close()


def build_report(
    flat_request: dict[str, str],
    sheet_row: tuple[Any, ...],
    monthly_blocks: list[MonthlyBlock],
    sheet_layout: SheetLayout,
    expense_details: dict[str, dict[str, float]] | None = None,
    occupant_name: str | None = None,
) -> dict[str, Any]:
    carry_over = safe_number(sheet_row[sheet_layout.carry_over_idx] if sheet_layout.carry_over_idx is not None and len(sheet_row) > sheet_layout.carry_over_idx else 0)
    total_collection = safe_number(sheet_row[sheet_layout.total_collection_idx] if sheet_layout.total_collection_idx is not None and len(sheet_row) > sheet_layout.total_collection_idx else 0)
    total_expense = safe_number(sheet_row[sheet_layout.total_expense_idx] if sheet_layout.total_expense_idx is not None and len(sheet_row) > sheet_layout.total_expense_idx else 0)
    total_late_fee = safe_number(sheet_row[sheet_layout.total_late_fee_idx] if sheet_layout.total_late_fee_idx is not None and len(sheet_row) > sheet_layout.total_late_fee_idx else 0)
    total_net = safe_number(sheet_row[sheet_layout.total_net_idx] if sheet_layout.total_net_idx is not None and len(sheet_row) > sheet_layout.total_net_idx else 0)
    total_dues = safe_number(sheet_row[sheet_layout.total_dues_idx] if sheet_layout.total_dues_idx is not None and len(sheet_row) > sheet_layout.total_dues_idx else 0)
    closing_balance = total_dues

    report = {
        "flat": flat_request["flat"],
        "owner_name": flat_request["owner_name"],
        "email": flat_request["email"],
        "sheet_email": (sheet_row[1] or "").strip() if len(sheet_row) > 1 and isinstance(sheet_row[1], str) else sheet_row[1],
        "occupant": occupant_name or None,
        "balance_from_last_fy": carry_over,
        "monthly": [],
        "totals": {},
    }

    expense_breakdown: list[dict[str, Any]] = []
    for block in monthly_blocks:
        exp_detail_total = expense_details.get(block.month_label, {}).get("total_expense") if (expense_details and block.month_label in expense_details) else None
        row_exp = safe_number(sheet_row[block.expense_idx] if len(sheet_row) > block.expense_idx else 0)
        expense_val = exp_detail_total if exp_detail_total is not None else row_exp

        exp_detail_late = expense_details.get(block.month_label, {}).get("late_fee") if (expense_details and block.month_label in expense_details) else None
        row_late = safe_number(sheet_row[block.late_fee_idx] if len(sheet_row) > block.late_fee_idx else 0)
        late_fee_val = exp_detail_late if exp_detail_late is not None else row_late

        exp_detail_reason = expense_details.get(block.month_label, {}).get("late_fee_reason") if (expense_details and block.month_label in expense_details) else None
        late_fee_reason = exp_detail_reason if (exp_detail_reason and str(exp_detail_reason).strip()) else "-"

        coll_val = safe_number(sheet_row[block.collection_idx] if len(sheet_row) > block.collection_idx else 0)
        net_val = coll_val - (expense_val + late_fee_val)

        report["monthly"].append(
            {
                "month": block.month_label,
                "collection": coll_val,
                "expense": expense_val,
                "late_fee": late_fee_val,
                "late_fee_reason": late_fee_reason,
                "net_collection_minus_expense": net_val,
            }
        )


        if expense_details and block.month_label in expense_details:
            expense_breakdown.append({"month": block.month_label, **expense_details[block.month_label]})
        else:
            expense_breakdown.append({"month": block.month_label})

    calc_total_expense = sum(m["expense"] for m in report["monthly"])
    calc_total_late = sum(m["late_fee"] for m in report["monthly"])
    calc_total_net = total_collection - (calc_total_expense + calc_total_late)
    closing_balance_calc = calc_total_net + carry_over

    report["totals"] = {
        "collection": total_collection,
        "expense": calc_total_expense if calc_total_expense > 0 else total_expense,
        "late_fee": calc_total_late if calc_total_late > 0 else total_late_fee,
        "net_collection_minus_expense": calc_total_net if calc_total_expense > 0 else total_net,
        "closing_balance": closing_balance_calc if calc_total_expense > 0 else closing_balance,
    }
    report["expense_breakdown"] = expense_breakdown
    return report


def generate_reports(
    workbook_path: Path,
    sheet_name: str,
    flat_requests: list[dict[str, str]],
    occupants: dict[str, str] | None = None,
 ) -> tuple[list[dict[str, Any]], list[str], str]:
    monthly_blocks, sheet_layout, row_lookup, available_sheets = load_sheet_rows(workbook_path, sheet_name)
    financial_year = derive_financial_year(monthly_blocks)

    if not has_cached_values(row_lookup, monthly_blocks):
        row_lookup = fill_from_source_sheets(
            workbook_path, monthly_blocks, sheet_layout, row_lookup, available_sheets
        )

    expense_sheet_map = build_expense_sheet_map(monthly_blocks, available_sheets)
    all_expense_details = load_expense_details(workbook_path, expense_sheet_map)
    occupants = occupants or {}
    reports: list[dict[str, Any]] = []
    missing_flats: list[str] = []

    for flat_request in flat_requests:
        flat = flat_request["flat"]
        sheet_row = row_lookup.get(flat)
        if sheet_row is None:
            missing_flats.append(flat)
            continue
        flat_expenses = all_expense_details.get(flat)
        occupant = occupants.get(flat, "")
        report = build_report(flat_request, sheet_row, monthly_blocks, sheet_layout, flat_expenses, occupant)
        reports.append(report)
    return reports, missing_flats, financial_year


def load_annual_expense_details(workbook_path: Path) -> list[dict[str, Any]]:
    """Load the ANNUAL-EXPENSE-DETAILS sheet: per-month rows with individual expense line items and summary columns."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_name = "ANNUAL-EXPENSE-DETAILS"
        if sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[sheet_name]
        
        summary_indices, col_names = get_aed_column_map(ws)

        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            month_name = row[0] if row[0] else None
            if not month_name:
                continue
            month_str = str(month_name).strip()
            entry: dict[str, Any] = {"month": month_str}

            # Individual line items (only include non-zero)
            line_items: dict[str, float] = {}
            for ci, name in col_names:
                val = safe_number(row[ci] if len(row) > ci else 0)
                if val:
                    line_items[name] = val
            entry["line_items"] = line_items

            # Summary columns
            for key in ["gross_expense", "gross_variable_expense", "gross_fixed_expense", "water_meter_rent", "total_expense"]:
                idx = summary_indices.get(key)
                entry[key] = safe_number(row[idx] if idx is not None and len(row) > idx else 0)

            rows.append(entry)
        return rows
    finally:
        workbook.close()


def main() -> int:
    workbooks_config: dict[str, Any] = json.loads(WORKBOOKS_JSON.read_text(encoding="utf-8"))
    portal_password = workbooks_config.get("portal_password")
    
    flat_requests = load_flat_requests(MEMBERS_CSV)
    occupants = load_occupants(OCCUPANTS_CSV)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    manifest_entries: list[dict[str, str]] = []
    for fy_key in sorted(workbooks_config.keys()):
        if fy_key == "portal_password":
            continue
            
        entry = workbooks_config[fy_key]
        workbook_path = PROJECT_ROOT / entry["workbook"]
        if not workbook_path.exists():
            print(f"Warning: workbook not found for FY {fy_key}: {workbook_path}")
            continue

        reports, missing_flats, financial_year = generate_reports(
            workbook_path=workbook_path,
            sheet_name=DEFAULT_SHEET,
            flat_requests=flat_requests,
            occupants=occupants,
        )
        annual_expense_details = load_annual_expense_details(workbook_path)
        cutoff_date = entry.get("cutoff-date-for-collection")
        fy_filename = f"report-data-{fy_key}.json"
        fy_output = OUTPUT_DIR / fy_filename
        write_report_dataset(reports, fy_output, financial_year, annual_expense_details, cutoff_date)
        manifest_entries.append({"fy": fy_key, "file": fy_filename})

        print(f"FY {fy_key}: generated {len(reports)} report(s) at {fy_output}")
        if missing_flats:
            print(f"  Missing flats: {', '.join(sorted(missing_flats))}")

    manifest = {"financial_years": manifest_entries}
    manifest_path = OUTPUT_DIR / "report-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"Manifest written to {manifest_path}")
    
    if portal_password:
        update_portal_password(portal_password)
        
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
