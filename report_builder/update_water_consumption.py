"""
Water Consumption Update Script

Overview:
---------
This script automates updating flat-wise water consumption readings in the main
society financial workbook for a given month using raw WaterOn consumption reports.

Pre-conditions:
---------------
1. Filename Convention:
   The input WaterOn consumption report filename MUST follow the pattern
   'Consumption Report <Month>-<YYYY>*.xlsx' (e.g. 'Consumption Report July-2026-ver1-8-10-26.xlsx').

2. Configuration Entry in workbooks.json:
   'db/workbooks.json' MUST exist and contain a mapping entry for the target Financial Year
   (e.g., "2026-27": {"workbook": "db/accounts/2026-2027.xlsx"}).

3. Target Workbook & Sheet Existence:
   - The target FY workbook Excel file MUST exist at the path configured in 'db/workbooks.json'.
   - The workbook MUST contain an expense sheet named '{MonthAbbr}{Year}-EXPENSE' or
     '{MonthFull}{Year}-EXPENSE' (e.g. 'Jul2026-EXPENSE' or 'July2026-EXPENSE').

4. Sheet Column Header Structure:
   Row 2 of the target expense sheet MUST contain a column header named 'WATER USED IN LTRS'.

5. WaterOn Report Structure:
   The first row of the WaterOn consumption Excel file MUST contain 'Apartment' and 'Total' columns.

6. File Lock / Write Access:
   The target financial workbook Excel file MUST NOT be open in Microsoft Excel or locked
   by another application.

Workflow:
---------
1. Parses the Month and Year from the input WaterOn consumption report filename
   (e.g., 'Consumption Report July-2026-ver1-8-10-26.xlsx').
2. Determines the Financial Year (April 1 to March 31, e.g. July 2026 -> FY '2026-27').
3. Resolves the target Excel workbook path via 'db/workbooks.json'.
4. Locates the monthly expense sheet (e.g., 'Jul2026-EXPENSE' or 'July2026-EXPENSE').
5. Dynamically finds the 'WATER USED IN LTRS' column header in row 2.
6. Reads total consumption per flat, normalizes flat identifiers, and populates cell values.

Usage:
------
python report_builder/update_water_consumption.py <path_to_wateron_report_excel>

Examples:
---------
1. Update July 2026 water consumption in FY 2026-27 workbook:
   python report_builder/update_water_consumption.py db/wateron/26-27/Consumption Report July-2026-ver1-8-10-26.xlsx

2. Run from anywhere within the repository:
   python report_builder/update_water_consumption.py "C:\github\manapristine\db\wateron\26-27\Consumption Report July-2026-ver1-8-10-26.xlsx"

Dependencies:
-------------
- openpyxl
- db/workbooks.json
"""

import os
import re
import json
import argparse
from datetime import datetime
from pathlib import Path
import openpyxl

def parse_wateron_filename(filepath):
    filename = os.path.basename(filepath)
    # Pattern: Consumption Report Month-YYYY.xlsx
    match = re.search(r'Consumption Report (?P<month>[a-zA-Z]+)-(?P<year>\d{4})', filename)
    if not match:
        raise ValueError(f"Could not parse month and year from filename: {filename}")
    
    month_str = match.group('month').lower()
    year = int(match.group('year'))
    
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    month = month_map.get(month_str)
    if not month:
        raise ValueError(f"Unknown month: {month_str}")
    
    return month, year

def get_financial_year(month, year):
    if month >= 4:
        return f"{year}-{str(year+1)[2:]}"
    else:
        return f"{year-1}-{str(year)[2:]}"

def normalize_flat(flat):
    if not flat:
        return ""
    flat = str(flat).strip().upper()
    mapping = {
        'C H': 'CH',
        'C GYM': 'GYM',
        'CBR': 'BSMT'
    }
    return mapping.get(flat, flat.replace(" ", ""))

def load_wateron_data(filepath):
    """Load water consumption data from WaterOn report."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active # Usually 'data'
        
        data = {}
        # Find columns in the first row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]
        
        try:
            apt_idx = headers.index('Apartment')
            total_idx = headers.index('Total')
        except ValueError:
            print("Error: Could not find 'Apartment' or 'Total' columns in WaterOn report.")
            return {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            apt = row[apt_idx]
            if apt:
                flat_id = normalize_flat(apt)
                if flat_id == "TOTAL":
                    continue
                total = row[total_idx]
                if isinstance(total, (int, float)):
                    data[flat_id] = total
        
        return data
    finally:
        if 'wb' in locals():
            wb.close()

def update_workbook(workbook_path, month, year, water_data):
    """Apply water consumption data to the corresponding workbook sheet."""
    print(f"Opening workbook: {workbook_path.name}")
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except PermissionError:
        print(f"Error: Permission denied. Please close the workbook '{workbook_path.name}' before running this script.")
        return False

    month_abbr = datetime(year, month, 1).strftime("%b")
    month_full = datetime(year, month, 1).strftime("%B")
    
    sheet_name = f"{month_abbr}{year}-EXPENSE"
    if sheet_name not in wb.sheetnames:
        sheet_name = f"{month_full}{year}-EXPENSE"
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet {month_abbr}{year}-EXPENSE or {month_full}{year}-EXPENSE not found.")
            return False

    ws = wb[sheet_name]
    print(f"Updating sheet: {sheet_name}")
    
    # Dynamically find columns in the header row (row 2)
    header_row = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[2]]
    
    try:
        flat_col_idx = 1 # Column A (1-based)
        water_col_idx = header_row.index('WATER USED IN LTRS') + 1
    except ValueError:
        print("Error: Could not find 'WATER USED IN LTRS' column in the workbook sheet.")
        return False

    updates = 0
    matched_flats = set()
    
    # Process rows starting from 3
    for row_idx in range(3, ws.max_row + 1):
        flat_cell = ws.cell(row=row_idx, column=flat_col_idx)
        flat_val = normalize_flat(flat_cell.value)
        
        if not flat_val or flat_val == "TOTAL":
            continue

        if flat_val in water_data:
            water_val = water_data[flat_val]
            cell = ws.cell(row=row_idx, column=water_col_idx)
            
            # Preserve formula warning (optional but good for production)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"  Warning: Overwriting formula in {flat_val} at row {row_idx}")
            
            cell.value = water_val
            updates += 1
            matched_flats.add(flat_val)

    # Reporting
    missing_in_wateron = set(water_data.keys()) - matched_flats
    if missing_in_wateron:
        print(f"\nNote: The following flats from the WaterOn report were NOT found in the workbook:")
        print(f"  {', '.join(sorted(missing_in_wateron))}")

    if updates > 0:
        try:
            wb.save(workbook_path)
            print(f"\nSuccessfully updated {updates} flats in {sheet_name}")
            return True
        except PermissionError:
            print(f"Error: Could not save workbook. Please ensure '{workbook_path.name}' is closed.")
            return False
    else:
        print("No matches found to update.")
        return False

def main():
    parser = argparse.ArgumentParser(description="Update water consumption from WaterOn report.")
    parser.add_argument('wateron_file', help="Path to the WaterOn consumption report Excel file.")
    args = parser.parse_args()
    
    if not os.path.exists(args.wateron_file):
        print(f"File not found: {args.wateron_file}")
        return

    try:
        month, year = parse_wateron_filename(args.wateron_file)
    except ValueError as e:
        print(e)
        return

    fy = get_financial_year(month, year)
    
    PROJECT_ROOT = Path(__file__).resolve().parent.parent
    WORKBOOKS_JSON = PROJECT_ROOT / 'db' / 'workbooks.json'
    
    if not WORKBOOKS_JSON.exists():
        print(f"Error: workbooks.json not found at {WORKBOOKS_JSON}")
        return
        
    with open(WORKBOOKS_JSON, 'r') as f:
        workbooks = json.load(f)
    
    if fy not in workbooks:
        print(f"Error: Financial year {fy} not found in workbooks.json")
        return
        
    workbook_path = PROJECT_ROOT / workbooks[fy]['workbook']
    if not workbook_path.exists():
        print(f"Error: Workbook not found: {workbook_path}")
        return

    water_data = load_wateron_data(args.wateron_file)
    if not water_data:
        print("No consumption data found in the WaterOn report.")
        return

    update_workbook(workbook_path, month, year, water_data)

if __name__ == "__main__":
    main()
