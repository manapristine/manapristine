#!/usr/bin/env python3
"""
Create a new financial year workbook from an existing one.

Copies the workbook, renames sheets for the new FY, updates all formula
references and text labels, shifts dates +1 year, and clears raw data
while preserving ALL formulas (including cross-sheet references).

Unlike refresh_fy.py, this script explicitly preserves formulas and only
zeros cells that contain raw numeric/date data.

Usage:
    python new_fy.py <workbook_path> <new_fy>

Arguments:
    workbook_path   Path to the source .xlsx workbook
    new_fy          Target financial year (e.g. "2027-28" means Apr 2027 to Mar 2028)

Example:
    python new_fy.py ../db/accounts/2026-2027-INCOME-EXPENDITURE-ACCOUNT-5.03.26-gold.xlsx 2027-28

Output:
    Creates a new file alongside the source with the new FY in the name.
"""

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


MONTHS_SHORT_ALT = ["Apr", "May", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_MIXED = [
    "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Jan", "Feb", "Mar", "June", "July",
]
APR_TO_DEC = {
    "full": ["APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"],
    "abbr": ["APR", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"],
    "mixed": ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "June", "July"],
}
JAN_TO_MAR = {
    "full": ["JANUARY", "FEBRUARY", "MARCH"],
    "abbr": ["JAN", "FEB", "MAR"],
    "mixed": ["Jan", "Feb", "Mar"],
}


def detect_fy_start(wb: openpyxl.Workbook) -> int:
    for name in wb.sheetnames:
        if name.startswith("Apr") and name.endswith("-EXPENSE"):
            mid = name.replace("Apr", "").replace("-EXPENSE", "")
            if mid.isdigit():
                return int(mid)
    raise ValueError("Cannot detect FY — no 'AprYYYY-EXPENSE' sheet found")


def parse_target_fy(fy_str: str) -> int:
    match = re.match(r"(\d{4})-(\d{2,4})", fy_str)
    if not match:
        raise ValueError(f"Invalid FY format '{fy_str}'. Expected YYYY-YY (e.g. 2027-28)")
    return int(match.group(1))


def build_rename_map(old_fy_start: int) -> dict[str, str]:
    new_apr = old_fy_start + 1
    old_jan = old_fy_start + 1
    new_jan = old_fy_start + 2

    renames: dict[str, str] = {}
    suffixes = ["-EXPENSE", "-COLLECTION"]

    for m in MONTHS_SHORT_ALT:
        for sfx in suffixes:
            renames[f"{m}{old_fy_start}{sfx}"] = f"{m}{new_apr}{sfx}"
    for sfx in suffixes:
        renames[f"June{old_fy_start}{sfx}"] = f"June{new_apr}{sfx}"
        renames[f"July{old_fy_start}{sfx}"] = f"July{new_apr}{sfx}"
    for extra in ["-EXPENSE-Flatwise", "-EXPENSE-Flatwise-Trans"]:
        renames[f"July{old_fy_start}{extra}"] = f"July{new_apr}{extra}"
    for m in ["Jan", "Feb", "Mar"]:
        for sfx in suffixes:
            renames[f"{m}{old_jan}{sfx}"] = f"{m}{new_jan}{sfx}"
    renames[f"INCOME-EXPENSE-CYCLES-{old_fy_start}-{old_fy_start + 1 - 2000}"] = (
        f"INCOME-EXPENSE-CYCLES-{new_apr}-{new_apr + 1 - 2000}"
    )
    return renames


def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def is_array_formula(v) -> bool:
    return isinstance(v, ArrayFormula)


def is_num(v) -> bool:
    return isinstance(v, (int, float))


def update_formula(text: str, sorted_renames: list[tuple[str, str]]) -> str:
    for old, new in sorted_renames:
        text = text.replace(f"'{old}'!", f"'{new}'!")
        text = text.replace(f"{old}!", f"{new}!")
    return text


def update_text(text: str, old_fy_start: int) -> str:
    if not isinstance(text, str):
        return text

    new_apr = old_fy_start + 1
    old_jan = old_fy_start + 1
    new_jan = old_fy_start + 2

    old_short = f"{old_fy_start}-{(old_fy_start + 1) % 100:02d}"
    new_short = f"{new_apr}-{(new_apr + 1) % 100:02d}"
    text = text.replace(f"Apr{old_fy_start} to Mar{old_jan}", f"Apr{new_apr} to Mar{new_jan}")
    text = text.replace(old_short, new_short)
    text = text.replace(f"{old_fy_start}-{old_fy_start + 1}", f"{new_apr}-{new_jan}")

    abbr_ms = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m in abbr_ms:
        text = text.replace(f"{m}'{(old_fy_start + 2) % 100:02d}", f"{m}'{(old_fy_start + 3) % 100:02d}")
    for m in abbr_ms:
        text = text.replace(f"{m}'{(old_fy_start + 1) % 100:02d}", f"{m}'{(old_fy_start + 2) % 100:02d}")
    for m in abbr_ms:
        text = text.replace(f"{m}'{old_fy_start % 100:02d}", f"{m}'{(old_fy_start + 1) % 100:02d}")

    for group in JAN_TO_MAR.values():
        for m in group:
            text = text.replace(f"{m} {old_jan}", f"{m} {new_jan}")
            text = text.replace(f"{m}{old_jan}", f"{m}{new_jan}")

    for group in APR_TO_DEC.values():
        for m in group:
            text = text.replace(f"{m} {old_fy_start}", f"{m} {new_apr}")
            text = text.replace(f"{m}{old_fy_start}", f"{m}{new_apr}")

    return text


def clear_collection_sheets(wb: openpyxl.Workbook, new_apr: int, new_jan: int) -> int:
    """Clear raw data in COLLECTION sheets. Preserve formulas (col K SUM) and flat/name cols."""
    cleared = 0
    sheet_names = (
        [f"{m}{new_apr}-COLLECTION" for m in MONTHS_SHORT_ALT]
        + [f"June{new_apr}-COLLECTION", f"July{new_apr}-COLLECTION"]
        + [f"{m}{new_jan}-COLLECTION" for m in ["Jan", "Feb", "Mar"]]
    )
    for sn in sheet_names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            # Cols 3-10: Amount/Date pairs (raw data)
            for c in [3, 5, 7, 9]:
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = None
                    cleared += 1
            for c in [4, 6, 8, 10]:
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (datetime, int, float)):
                    cell.value = None
                    cleared += 1
    return cleared


def clear_expense_sheets(wb: openpyxl.Workbook, new_apr: int, new_jan: int) -> int:
    """Clear raw data in EXPENSE sheets. Preserve all formulas."""
    cleared = 0
    sheet_names = (
        [f"{m}{new_apr}-EXPENSE" for m in MONTHS_SHORT_ALT]
        + [f"June{new_apr}-EXPENSE", f"July{new_apr}-EXPENSE"]
        + [f"{m}{new_jan}-EXPENSE" for m in ["Jan", "Feb", "Mar"]]
    )
    for sn in sheet_names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            # Col 3: water used (raw), Col 4: common area water (formula in some, raw in others)
            for c in [3, 4]:
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    cleared += 1
            # Cols 12-17: parking, club house, shifting, gym, covid, membership (raw data)
            for c in range(12, 18):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    cleared += 1
    return cleared


def clear_income_expense_cycles(wb: openpyxl.Workbook) -> int:
    """Clear raw data in INCOME-EXPENSE-CYCLES. NEVER touch formula cells."""
    cleared = 0
    if "INCOME-EXPENSE-CYCLES" not in wb.sheetnames:
        return 0
    ws = wb["INCOME-EXPENSE-CYCLES"]
    for r in range(3, ws.max_row + 1):
        # Col 4: Balance from last FY (numeric, manually entered)
        cell = ws.cell(row=r, column=4)
        if is_num(cell.value):
            cell.value = 0
            cleared += 1
        # Cols 5-52: monthly data — ONLY clear if numeric (NOT formula)
        for c in range(5, 53):
            cell = ws.cell(row=r, column=c)
            if is_num(cell.value) and not is_formula(cell.value):
                cell.value = 0
                cleared += 1
    return cleared


def clear_income_expense_variants(wb: openpyxl.Workbook, new_apr: int) -> int:
    """Clear data in variant INCOME-EXPENSE-CYCLES sheets."""
    cleared = 0
    fy_tag = f"{new_apr}-{(new_apr + 1) % 100:02d}"
    for sn in [f"INCOME-EXPENSE-CYCLES-{fy_tag}", "Copy of INCOME-EXPENSE-CYCLES 1", "Copy of INCOME-EXPENSE-CYCLES"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            for c in range(4, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value) and not is_formula(cell.value):
                    cell.value = 0
                    cleared += 1
    return cleared


def clear_annual_expense_details(wb: openpyxl.Workbook) -> int:
    """Clear numeric data in ANNUAL-EXPENSE-DETAILS rows 3-14."""
    cleared = 0
    if "ANNUAL-EXPENSE-DETAILS" not in wb.sheetnames:
        return 0
    ws = wb["ANNUAL-EXPENSE-DETAILS"]
    for r in range(3, 15):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if is_num(cell.value) and not is_formula(cell.value):
                cell.value = 0
                cleared += 1
    return cleared


def clear_flatwise_sheets(wb: openpyxl.Workbook, new_apr: int) -> int:
    """Clear July Flatwise sheets."""
    cleared = 0
    for sn in [f"July{new_apr}-EXPENSE-Flatwise", f"July{new_apr}-EXPENSE-Flatwise-Trans"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value) and not is_formula(cell.value) and not is_array_formula(cell.value):
                    cell.value = 0
                    cleared += 1
    return cleared


def clear_utility_sheets(wb: openpyxl.Workbook) -> int:
    """Clear utility/archive sheets."""
    cleared = 0
    for sn, start_col in [("Sheet30", 3), ("Sheet4", 3), ("Sheet27", 2)]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value) and not is_formula(cell.value):
                    cell.value = 0
                    cleared += 1
    if "2013-till-date-connected" in wb.sheetnames:
        ws = wb["2013-till-date-connected"]
        for r in range(4, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value) and not is_formula(cell.value):
                    cell.value = 0
                    cleared += 1
    return cleared


def refresh(source_path: str, target_fy_start: int) -> str:
    wb = openpyxl.load_workbook(source_path)
    old_fy_start = detect_fy_start(wb)
    new_apr = old_fy_start + 1
    new_jan = old_fy_start + 2

    if new_apr != target_fy_start:
        print(f"ERROR: Source workbook is FY {old_fy_start}-{(old_fy_start+1)%100:02d}.")
        print(f"  Next FY would be {new_apr}-{(new_apr+1)%100:02d}, but you requested {target_fy_start}-{(target_fy_start+1)%100:02d}.")
        print(f"  Can only advance by one year at a time.")
        sys.exit(1)

    print(f"Source: {source_path}")
    print(f"Refreshing FY {old_fy_start}-{(old_fy_start+1)%100:02d} → {new_apr}-{(new_apr+1)%100:02d}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print()

    # Step 1: Rename sheets
    rename_map = build_rename_map(old_fy_start)
    renamed = 0
    for old, new in rename_map.items():
        if old in wb.sheetnames:
            wb[old].title = new
            renamed += 1
    print(f"[1/6] Renamed {renamed} sheets")

    # Step 2: Update Members!F2 (FY_START_YEAR) — drives all INDIRECT formulas
    if "Members" in wb.sheetnames:
        wb["Members"].cell(2, 6, new_apr)
        print(f"[2/7] Updated Members!F2 = {new_apr} (drives INDIRECT cross-sheet refs)")
    else:
        print(f"[2/7] SKIP: No Members sheet found")

    # Step 3: Update formulas and text labels
    sorted_renames = sorted(rename_map.items(), key=lambda x: len(x[0]), reverse=True)
    updated = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if is_array_formula(v):
                    v.text = update_formula(v.text, sorted_renames)
                    updated += 1
                elif isinstance(v, str):
                    if v.startswith("="):
                        new_v = update_formula(v, sorted_renames)
                        new_v = update_text(new_v, old_fy_start)
                        if new_v != v:
                            cell.value = new_v
                            updated += 1
                    else:
                        new_v = update_text(v, old_fy_start)
                        if new_v != v:
                            cell.value = new_v
                            updated += 1
                elif isinstance(v, datetime):
                    try:
                        cell.value = v.replace(year=v.year + 1)
                    except ValueError:
                        cell.value = v.replace(year=v.year + 1, day=28)
                    updated += 1
    print(f"[3/7] Updated {updated} cells (formulas, labels, dates)")

    # Step 4-7: Clear raw data (preserving formulas)
    c1 = clear_collection_sheets(wb, new_apr, new_jan)
    print(f"[4/7] Cleared {c1} cells in COLLECTION sheets")

    c2 = clear_expense_sheets(wb, new_apr, new_jan)
    print(f"[5/7] Cleared {c2} cells in EXPENSE sheets")

    c3 = clear_income_expense_cycles(wb)
    c3 += clear_income_expense_variants(wb, new_apr)
    print(f"[6/7] Cleared {c3} cells in INCOME-EXPENSE-CYCLES sheets")

    c4 = clear_annual_expense_details(wb)
    c4 += clear_flatwise_sheets(wb, new_apr)
    c4 += clear_utility_sheets(wb)
    print(f"[7/7] Cleared {c4} cells in other sheets")

    # Verify: count formulas preserved
    formula_count = 0
    cross_sheet_count = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if is_formula(cell.value):
                    formula_count += 1
                    if "!" in cell.value:
                        cross_sheet_count += 1
                elif is_array_formula(cell.value):
                    formula_count += 1

    print(f"\nFormulas preserved: {formula_count} (including {cross_sheet_count} cross-sheet references)")

    # Save
    source_p = Path(source_path)
    new_filename = re.sub(
        r"\d{4}-\d{4}",
        f"{new_apr}-{new_jan}",
        source_p.name,
    )
    if new_filename == source_p.name:
        new_filename = f"{new_apr}-{new_jan}-{source_p.name}"
    output_path = source_p.parent / new_filename
    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Create a new financial year workbook from an existing one, preserving all formulas"
    )
    parser.add_argument("workbook", help="Path to the source .xlsx workbook")
    parser.add_argument("new_fy", help="Target financial year (e.g. '2027-28')")
    parser.add_argument("--no-backup", action="store_true", help="Skip creating a backup of the source")
    args = parser.parse_args()

    source = Path(args.workbook)
    if not source.exists():
        print(f"ERROR: File not found: {source}")
        sys.exit(1)

    target_fy_start = parse_target_fy(args.new_fy)

    if not args.no_backup:
        backup = str(source).replace(".xlsx", "-backup.xlsx")
        shutil.copy2(str(source), backup)
        print(f"Backup: {backup}\n")

    output_path = refresh(str(source), target_fy_start)

    print("\nDone. Next steps:")
    print(f"  1. Open {output_path} in Excel/Sheets and verify formulas calculate")
    print(f"  2. Update db/workbooks.json to register the new FY")
    print(f"  3. Run: python compare_workbooks.py {args.workbook} {output_path}")


if __name__ == "__main__":
    main()
