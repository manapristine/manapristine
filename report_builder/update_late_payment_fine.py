"""
Late Payment Fine Update Script

Overview:
---------
This script automates the calculation and application of late payment fines in society
financial workbooks. It inspects flat-wise maintenance payments in the immediately preceding
month collection sheet ('<prev_month>-COLLECTION') and updates the 'LATE PAYMENT FINE' column in
the current month expense sheet ('<curr_month>-EXPENSE').

It also verifies and updates the Excel formula in the 'TOTAL EXPENSE TO BE PAID' column across
all expense sub-sheets so that 'LATE PAYMENT FINE' is fully included alongside all other expense heads.

Rule & Policy:
--------------
- 1-Month Look Back: Evaluates payment status in the immediately preceding month.
- Fixed Fine: If payment was missed in the previous month's COLLECTION sheet AND the flat has net
  outstanding dues (cumulative collections < cumulative expenses up to that month), a fine of
  Rs 1,000 (or configured fine amount) is applied in the current month's EXPENSE sheet.
- Excess Payment Fine Waiver: If payment was missed in the previous month BUT the flat has an excess
  amount paid (cumulative collections >= cumulative expenses up to that month, including opening
  balance from last FY), the late payment fine is WAIVED (Rs 0).
- On-Time Payment: If payment was received in the previous month, the fine is Rs 0.

Pre-conditions:
---------------
1. Workbook Structure:
   The target society financial workbook MUST exist and contain paired monthly collection and
   expense sheets (e.g., 'Apr2026-COLLECTION' and 'May2026-EXPENSE').

2. Collection Sheet Structure:
   Each '<month>-COLLECTION' sheet MUST contain flat numbers in Column A (1) and payment amount cells
   (Columns 3, 5, 7, 9) / 'Total' column indicating the total maintenance received for that flat in that month.

3. Expense Sheet Structure:
   Each '<month>-EXPENSE' sheet header row (Row 2) MUST contain a column named 'LATE PAYMENT FINE'
   and a column named 'TOTAL EXPENSE TO BE PAID'.

4. File Lock / Write Access:
   The target society financial workbook MUST NOT be open in Microsoft Excel or locked by another application.

Key Configurations & Assumptions:
---------------------------------
- DEFAULT_FINE_PER_MONTH: Rs 1,000 per month of non-payment.
- Unpaid Determination: A flat is considered unpaid in a collection month if its 'Total' cell value
  and payment amount cells (Columns C, E, G, I) are None, empty, or 0.

CLI Usage:
----------
1. Basic Usage (uses default active workbook from db/workbooks.json):
   python report_builder/update_late_payment_fine.py

2. Explicit Workbook Path:
   python report_builder/update_late_payment_fine.py "db/accounts/2026-2027-INCOME-EXPENDITURE-ACCOUNT-8.10.26-gold.xlsx"

3. Custom Fine Amount (e.g. Rs 500 per month):
   python report_builder/update_late_payment_fine.py "path/to/workbook.xlsx" --fine 500

4. Help Menu:
   python report_builder/update_late_payment_fine.py --help

Exit Codes:
-----------
  0: Success (workbook updated successfully)
  1: Failure (missing file, missing sheet/column, permission error)
"""

import os
import re
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter

# Project Root and Default Configuration
PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOKS_JSON = PROJECT_ROOT / "db" / "workbooks.json"
DEFAULT_FINE_PER_MONTH = 1000  # Rs 1,000 flat fine for missed payment in previous month

def normalize_flat(flat):
    """Normalize flat identifiers to consistent uppercase standard (e.g. F1, G16, CH)."""
    if flat is None:
        return ""
    flat = str(flat).strip().upper()
    mapping = {
        'C H': 'CH',
        'C GYM': 'GYM',
        'CBR': 'BSMT'
    }
    flat = mapping.get(flat, flat.replace(" ", ""))
    flat = re.sub(r'^([A-Z])0+(\d+)$', r'\1\2', flat)
    return flat

def get_active_workbook_path():
    """Retrieve default active financial workbook path from db/workbooks.json."""
    if not WORKBOOKS_JSON.exists():
        print(f"Error: workbooks.json not found at {WORKBOOKS_JSON}")
        return None
    with open(WORKBOOKS_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    for fy_key in ["2026-27", "2025-26"]:
        if fy_key in data and "workbook" in data[fy_key]:
            wb_rel = data[fy_key]["workbook"]
            wb_path = PROJECT_ROOT / wb_rel
            if wb_path.exists():
                return wb_path
    
    for fy_key, cfg in data.items():
        if isinstance(cfg, dict) and "workbook" in cfg:
            wb_path = PROJECT_ROOT / cfg["workbook"]
            if wb_path.exists():
                return wb_path
    return None

def find_sheet(wb, candidates):
    """Find sheet name in workbook matching candidate list."""
    for cand in candidates:
        if cand in wb.sheetnames:
            return cand
    return None

def parse_fy_months_from_sheetnames(sheetnames):
    """Detect start year and return full 12-month FY sequence (April to March) as (month, year) tuples."""
    years = sorted({int(y) for s in sheetnames for y in re.findall(r'\d{4}', s)})
    if not years:
        start_year = datetime.now().year
    else:
        start_year = years[0]
    
    seq = [(m, start_year) for m in range(4, 13)] + [(m, start_year + 1) for m in range(1, 4)]
    return seq

def is_flat_paid_in_collection_row(ws_c, row_idx, total_col_idx):
    """Determine if maintenance payment was received for flat in collection sheet row."""
    total_val = ws_c.cell(row=row_idx, column=total_col_idx).value
    
    # 1. Check direct numeric total cell value if available
    if isinstance(total_val, (int, float)):
        return total_val > 0

    # 2. If Total cell contains formula string or None, inspect payment amount columns (3, 5, 7, 9)
    pmt_sum = 0.0
    for col_idx in (3, 5, 7, 9):
        val = ws_c.cell(row=row_idx, column=col_idx).value
        if isinstance(val, (int, float)) and val > 0:
            pmt_sum += float(val)
        elif isinstance(val, str):
            clean_str = re.sub(r'[^0-9.]', '', val.strip())
            if clean_str:
                try:
                    num = float(clean_str)
                    if num > 0:
                        pmt_sum += num
                except ValueError:
                    pass
    
    return pmt_sum > 0

def ensure_total_expense_formulas(wb):
    """
    Ensure the formula in 'TOTAL EXPENSE TO BE PAID' column includes all expense heads
    (from Column 10/J up to the LATE PAYMENT FINE column) across all EXPENSE sheets.

    :param wb: openpyxl Workbook object
    :return: int number of formula cells updated
    """
    formula_updates = 0
    non_flat_ids = {"TOTAL", "CH", "GYM", "BSMT", "MPFOWA"}

    for s in wb.sheetnames:
        if 'EXPENSE' in s and s != 'ANNUAL-EXPENSE-DETAILS':
            ws = wb[s]
            headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            
            total_col_idx = None
            for idx, h in enumerate(headers, start=1):
                if h and 'TOTAL EXPENSE TO BE PAID' in str(h).upper():
                    total_col_idx = idx
                    break
            
            if total_col_idx and total_col_idx > 10:
                last_head_col_letter = get_column_letter(total_col_idx - 1)
                
                for r in range(3, ws.max_row + 1):
                    flat_val = ws.cell(r, 1).value
                    flat_id = normalize_flat(flat_val)
                    if not flat_id or flat_id in non_flat_ids:
                        continue

                    # Build Excel sum formula: =SUM(J{r}:{last_head_col_letter}{r})*$B$2
                    target_formula = f"=SUM(J{r}:{last_head_col_letter}{r})*$B$2"
                    curr_formula = str(ws.cell(r, total_col_idx).value or "").strip().upper()

                    if curr_formula != target_formula.upper():
                        ws.cell(r, total_col_idx).value = target_formula
                        formula_updates += 1

    return formula_updates

def update_late_payment_fines(workbook_path, fine_per_month=DEFAULT_FINE_PER_MONTH, as_of_date=None):
    """
    Inspect immediately preceding month collection sheet and update LATE PAYMENT FINE in expense sheets.
    Does NOT update late payment fines for the current month or future months beyond as_of_date,
    as collection amounts received in current month are for the previous month.
    Waives late payment fine if flat has excess amount paid (cumulative collections >= cumulative expenses).
    Also ensures formulas in TOTAL EXPENSE TO BE PAID include LATE PAYMENT FINE across all expense sub-sheets.

    :param workbook_path: Path to Excel financial workbook.
    :param fine_per_month: Late fee for missed payment in previous month (default: Rs 1000).
    :param as_of_date: Reference cutoff date string in 'YYYY-MM' format (default: current month).
    :return: bool success
    """
    wb_path = Path(workbook_path)
    if not wb_path.exists() and not wb_path.is_absolute():
        alt_path = PROJECT_ROOT / workbook_path
        if alt_path.exists():
            wb_path = alt_path

    if not wb_path.exists():
        print(f"Error: Workbook file not found: {wb_path}")
        return False

    if as_of_date:
        try:
            dt = datetime.strptime(as_of_date, "%Y-%m")
            as_of_yr, as_of_m = dt.year, dt.month
        except ValueError:
            print(f"Error: Invalid --as-of date format '{as_of_date}'. Expected YYYY-MM.")
            return False
    else:
        now = datetime.now()
        as_of_yr, as_of_m = now.year, now.month

    print(f"Opening workbook: {wb_path.name}")
    print(f"Late fine calculation cutoff month: {as_of_yr}-{as_of_m:02d} (Current & future months skipped)")

    try:
        wb = openpyxl.load_workbook(wb_path)
        wb_data = openpyxl.load_workbook(wb_path, data_only=True)
    except PermissionError:
        print(f"Error: Permission denied opening '{wb_path.name}'. Ensure file is closed in Excel.")
        return False

    month_seq = parse_fy_months_from_sheetnames(wb.sheetnames)
    updates_total = 0
    sheet_summaries = []
    non_flat_ids = {"TOTAL", "CH", "GYM", "BSMT", "MPFOWA"}

    # Load carryover balance from INCOME-EXPENSE-CYCLES ('Balance from last FY year')
    carryover_balances = {}
    if 'INCOME-EXPENSE-CYCLES' in wb_data.sheetnames:
        ws_iec = wb_data['INCOME-EXPENSE-CYCLES']
        row1 = [ws_iec.cell(1, col).value for col in range(1, ws_iec.max_column + 1)]
        bal_col_idx = 4
        for idx_c, val in enumerate(row1, start=1):
            if val and 'BALANCE FROM LAST FY YEAR' in str(val).upper():
                bal_col_idx = idx_c
                break
        for r in range(3, ws_iec.max_row + 1):
            flat_val = ws_iec.cell(row=r, column=1).value
            flat_id = normalize_flat(flat_val)
            if not flat_id or flat_id in non_flat_ids:
                continue
            bal_val = ws_iec.cell(row=r, column=bal_col_idx).value
            if isinstance(bal_val, (int, float)):
                carryover_balances[flat_id] = float(bal_val)

    cum_collections = defaultdict(float, {f: carryover_balances.get(f, 0.0) for f in carryover_balances})
    cum_expenses = defaultdict(float)

    # Pre-read all monthly collection and expense totals from wb_data
    monthly_colls = defaultdict(lambda: defaultdict(float))
    monthly_exps = defaultdict(lambda: defaultdict(float))

    for m, yr in month_seq:
        dt_m = datetime(yr, m, 1)
        m_abbr, m_full = dt_m.strftime('%b'), dt_m.strftime('%B')
        coll_sname = find_sheet(wb_data, [f"{m_abbr}{yr}-COLLECTION", f"{m_full}{yr}-COLLECTION"])
        exp_sname = find_sheet(wb_data, [f"{m_abbr}{yr}-EXPENSE", f"{m_full}{yr}-EXPENSE"])

        if coll_sname:
            ws_c = wb_data[coll_sname]
            t_col = 11
            header_c = [str(ws_c.cell(1, c).value).strip().upper() if ws_c.cell(1, c).value else "" for c in range(1, ws_c.max_column + 1)]
            if "TOTAL" in header_c:
                t_col = header_c.index("TOTAL") + 1
            for r in range(2, ws_c.max_row + 1):
                fid = normalize_flat(ws_c.cell(r, 1).value)
                if fid and fid not in non_flat_ids:
                    amt = 0.0
                    t_val = ws_c.cell(r, t_col).value
                    if isinstance(t_val, (int, float)):
                        amt = float(t_val)
                    else:
                        for c_idx in (3, 5, 7, 9):
                            v = ws_c.cell(r, c_idx).value
                            if isinstance(v, (int, float)):
                                amt += float(v)
                    monthly_colls[(m, yr)][fid] = amt

        if exp_sname:
            ws_e = wb_data[exp_sname]
            tot_exp_col = 19
            header_e = [str(ws_e.cell(2, c).value).strip().upper() if ws_e.cell(2, c).value else "" for c in range(1, ws_e.max_column + 1)]
            for c_idx, h in enumerate(header_e, start=1):
                if "TOTAL EXPENSE TO BE PAID" in h:
                    tot_exp_col = c_idx
                    break

            for r in range(3, ws_e.max_row + 1):
                fid = normalize_flat(ws_e.cell(r, 1).value)
                if fid and fid not in non_flat_ids:
                    e_val = ws_e.cell(r, tot_exp_col).value
                    if e_val is None or not isinstance(e_val, (int, float)):
                        raise ValueError(
                            f"Null or missing cached formula value for 'TOTAL EXPENSE TO BE PAID' in workbook '{wb_path.name}', "
                            f"sheet '{exp_sname}', flat '{fid}'. Please open the workbook in Excel, save it, close it, and re-run."
                        )
                    monthly_exps[(m, yr)][fid] = float(e_val)

    try:
        # Iterate sequentially through FY months
        for idx in range(len(month_seq)):
            curr_m, curr_yr = month_seq[idx]
            dt_curr = datetime(curr_yr, curr_m, 1)
            curr_abbr = dt_curr.strftime('%b')
            curr_full = dt_curr.strftime('%B')

            collection_sheet_name = find_sheet(wb_data, [f"{curr_abbr}{curr_yr}-COLLECTION", f"{curr_full}{curr_yr}-COLLECTION"])
            expense_sheet_name = find_sheet(wb_data, [f"{curr_abbr}{curr_yr}-EXPENSE", f"{curr_full}{curr_yr}-EXPENSE"])

            # 1. Process late fee application for current month's EXPENSE sheet (looks back at prev_m)
            if idx > 0 and expense_sheet_name and expense_sheet_name in wb.sheetnames:
                prev_m, prev_yr = month_seq[idx - 1]
                dt_prev = datetime(prev_yr, prev_m, 1)
                prev_abbr = dt_prev.strftime('%b')
                prev_full = dt_prev.strftime('%B')
                prev_collection_sheet_name = find_sheet(wb_data, [f"{prev_abbr}{prev_yr}-COLLECTION", f"{prev_full}{prev_yr}-COLLECTION"])

                is_current = (curr_yr, curr_m) == (as_of_yr, as_of_m)
                is_future = (curr_yr, curr_m) > (as_of_yr, as_of_m)
                skip_late_fee = is_current or is_future

                prev_month_missed = {}
                has_collection_data = False

                if prev_collection_sheet_name and not skip_late_fee:
                    ws_pc = wb_data[prev_collection_sheet_name]
                    total_col_idx = 11
                    header_c = [str(ws_pc.cell(1, col).value).strip().upper() if ws_pc.cell(1, col).value else "" for col in range(1, ws_pc.max_column + 1)]
                    if "TOTAL" in header_c:
                        total_col_idx = header_c.index("TOTAL") + 1

                    for r in range(2, ws_pc.max_row + 1):
                        flat_val = ws_pc.cell(row=r, column=1).value
                        flat_id = normalize_flat(flat_val)
                        if not flat_id or flat_id in non_flat_ids:
                            continue

                        paid = is_flat_paid_in_collection_row(ws_pc, r, total_col_idx)
                        prev_month_missed[flat_id] = not paid
                        if paid:
                            has_collection_data = True

                ws_e = wb[expense_sheet_name]
                header_e = [str(ws_e.cell(2, col).value).strip().upper() if ws_e.cell(2, col).value else "" for col in range(1, ws_e.max_column + 1)]

                if "LATE PAYMENT FINE" in header_e:
                    fine_col_idx = header_e.index("LATE PAYMENT FINE") + 1
                    sheet_updates = 0
                    flagged_in_sheet = []
                    waived_in_sheet = []

                    for r in range(3, ws_e.max_row + 1):
                        flat_val = ws_e.cell(row=r, column=1).value
                        flat_id = normalize_flat(flat_val)
                        if not flat_id or flat_id in non_flat_ids:
                            continue

                        if skip_late_fee or not has_collection_data:
                            fine_amount = 0
                        else:
                            missed = prev_month_missed.get(flat_id, False)
                            if missed:
                                net_pos = cum_collections[flat_id] - cum_expenses[flat_id]
                                if net_pos >= 0:
                                    fine_amount = 0
                                    waived_in_sheet.append((flat_id, net_pos))
                                else:
                                    fine_amount = fine_per_month
                                    flagged_in_sheet.append((flat_id, fine_amount, net_pos))
                            else:
                                fine_amount = 0

                        fine_cell = ws_e.cell(row=r, column=fine_col_idx)
                        existing_val = fine_cell.value
                        try:
                            clean_str = re.sub(r'[^0-9.]', '', str(existing_val)) if existing_val is not None else ""
                            existing_num = float(clean_str) if clean_str else 0.0
                        except (ValueError, TypeError):
                            existing_num = -1.0

                        if existing_num != float(fine_amount):
                            fine_cell.value = fine_amount
                            sheet_updates += 1
                            updates_total += 1

                    sheet_summaries.append((expense_sheet_name, prev_collection_sheet_name, sheet_updates, flagged_in_sheet, waived_in_sheet, is_current, is_future, has_collection_data))

            # 2. Accumulate current month collections & expenses into running cumulative totals for subsequent months
            all_flats = set(list(monthly_colls[(curr_m, curr_yr)].keys()) + list(monthly_exps[(curr_m, curr_yr)].keys()))
            for fid in all_flats:
                cum_collections[fid] += monthly_colls[(curr_m, curr_yr)][fid]
                cum_expenses[fid] += monthly_exps[(curr_m, curr_yr)][fid]

        wb_data.close()

        # 4. Ensure all TOTAL EXPENSE TO BE PAID formulas include LATE PAYMENT FINE
        formula_updates = ensure_total_expense_formulas(wb)

        if updates_total > 0 or formula_updates > 0 or sheet_summaries:
            try:
                wb.save(wb_path)
                print(f"\nSuccessfully saved updated workbook: '{wb_path.name}'")
                print(f"Total Late Payment Fine updates applied: {updates_total}")
                print(f"Total EXPENSE formula updates applied: {formula_updates}\n")

                for sheet_name, col_name, count, flagged, waived, is_curr, is_fut, has_data in sheet_summaries:
                    print(f"--- Sheet: {sheet_name} (Based on {col_name or 'N/A'}) ---")
                    if is_curr:
                        print("  Skipped: Current month (collections still underway for previous month).")
                    elif is_fut:
                        print("  Skipped: Future month beyond cutoff date.")
                    elif not has_data:
                        print("  Skipped: Previous month has no recorded collection data.")
                    else:
                        if flagged:
                            print(f"  {'Flat':<8} | {'Payment Status':<20} | {'Net Position (Rs)':<20} | {'Late Fine (Rs)':<15}")
                            print("  " + "-" * 70)
                            for fid, fine, net in flagged:
                                print(f"  {fid:<8} | {'MISSED':<20} | Rs {net:<17,f} | Rs {fine:<13,d}")
                        if waived:
                            print(f"\n  [WAIVED - Excess Payment Credit Available]")
                            print(f"  {'Flat':<8} | {'Payment Status':<20} | {'Net Position (Rs)':<20} | {'Late Fine (Rs)':<15}")
                            print("  " + "-" * 70)
                            for fid, net in waived:
                                print(f"  {fid:<8} | {'MISSED':<20} | +Rs {net:<16,f} | Rs 0 (WAIVED)")
                        if not flagged and not waived:
                            print("  No late payment fines applicable (All flats paid on time in previous month).")
                    print()
                return True
            except PermissionError:
                print(f"Error: Could not save '{wb_path.name}'. File is locked/open in Excel.")
                return False
        else:
            print("\nNo updates were required.")
            return True
    finally:
        wb.close()


def main():
    parser = argparse.ArgumentParser(description="Update LATE PAYMENT FINE column in expense sheets based on 1-month lookback collection status.")
    parser.add_argument(
        "accounts_file",
        nargs="?",
        default=None,
        help="Path to the society financial accounts Excel workbook file."
    )
    parser.add_argument(
        "--fine", "-f",
        type=int,
        default=DEFAULT_FINE_PER_MONTH,
        help=f"Late payment fine amount for missed payment in previous month in Rupees (default: Rs {DEFAULT_FINE_PER_MONTH})."
    )
    parser.add_argument(
        "--as-of", "-a",
        default=None,
        help="Reference cutoff month in YYYY-MM format (e.g. 2026-08). Defaults to current month."
    )
    args = parser.parse_args()

    wb_path = args.accounts_file
    if wb_path is None:
        wb_path = get_active_workbook_path()
        if wb_path is None:
            print("Error: Could not resolve default workbook path from db/workbooks.json")
            sys.exit(1)

    success = update_late_payment_fines(wb_path, fine_per_month=args.fine, as_of_date=args.as_of)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
