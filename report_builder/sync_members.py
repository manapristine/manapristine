#!/usr/bin/env python3
"""
Sync db/members.csv into the Members sheet of all registered workbooks.

Reads members.csv and updates the Members sheet in each workbook listed
in db/workbooks.json. Run this after editing members.csv.

Usage:
    python sync_members.py

Or to sync a specific workbook:
    python sync_members.py --workbook ../db/accounts/some-workbook.xlsx
"""

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent.parent
MEMBERS_CSV = PROJECT_ROOT / "db" / "members.csv"
WORKBOOKS_JSON = PROJECT_ROOT / "db" / "workbooks.json"


def load_members() -> list[dict[str, str]]:
    members = []
    with MEMBERS_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            members.append({
                "flat": row.get("flat", "").strip().upper(),
                "name": row.get("name", "").strip(),
                "email": row.get("email", "").strip(),
                "phone": row.get("phone", "").strip(),
            })
    return members


def sync_workbook(workbook_path: Path, members: list[dict[str, str]]) -> bool:
    if not workbook_path.exists():
        print(f"  SKIP: {workbook_path} (not found)")
        return False

    wb = openpyxl.load_workbook(str(workbook_path))

    if "Members" not in wb.sheetnames:
        ws = wb.create_sheet("Members", 0)
        ws.cell(1, 1, "FLAT")
        ws.cell(1, 2, "NAME")
        ws.cell(1, 3, "EMAIL")
        ws.cell(1, 4, "PHONE")
        print(f"  Created new Members sheet")
    else:
        ws = wb["Members"]

    # Update data rows
    for i, m in enumerate(members):
        row = i + 2
        ws.cell(row, 1, m["flat"])
        ws.cell(row, 2, m["name"])
        ws.cell(row, 3, m["email"])
        ws.cell(row, 4, m["phone"])

    # Clear any extra rows beyond the current member count
    for row in range(len(members) + 2, ws.max_row + 1):
        for col in range(1, 5):
            ws.cell(row, col, None)

    wb.save(str(workbook_path))
    print(f"  Synced {len(members)} members into {workbook_path.name}")
    return True


def main():
    parser = argparse.ArgumentParser(description="Sync members.csv into workbook Members sheets")
    parser.add_argument("--workbook", help="Path to a specific workbook (otherwise syncs all in workbooks.json)")
    args = parser.parse_args()

    members = load_members()
    print(f"Loaded {len(members)} members from {MEMBERS_CSV.name}")

    if args.workbook:
        path = Path(args.workbook)
        if not path.is_absolute():
            path = PROJECT_ROOT / path
        sync_workbook(path, members)
    else:
        config = json.loads(WORKBOOKS_JSON.read_text(encoding="utf-8"))
        latest_fy = sorted(config.keys())[-1]
        entry = config[latest_fy]
        workbook_path = PROJECT_ROOT / entry["workbook"]
        print(f"\nFY {latest_fy} (latest):")
        sync_workbook(workbook_path, members)
        print("\nNote: Only the latest FY workbook is synced by default.")
        print("Use --workbook to sync a specific file.")

    print("\nDone. Open the workbook(s) in Excel/Sheets and save to refresh formula cache.")


if __name__ == "__main__":
    main()
