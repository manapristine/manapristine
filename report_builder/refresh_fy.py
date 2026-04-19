#!/usr/bin/env python3
"""
Refresh the INCOME-EXPENDITURE-ACCOUNT workbook for a new financial year.

Renames sheets, updates formula references, shifts dates by +1 year,
updates text labels, and zeroes out raw data — while preserving all formulas.

Usage:
    python refresh_fy.py <workbook_path>

Example:
    python refresh_fy.py ../db/accounts/2026-2027-INCOME-EXPENDITURE-ACCOUNT-4.19.26.xlsx
"""

import argparse
import shutil
import sys
from datetime import datetime

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula


MONTHS_SHORT = ["Apr", "May", "June", "July", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
MONTHS_SHORT_ALT = ["Apr", "May", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_FULL = [
    "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER",
    "OCTOBER", "NOVEMBER", "DECEMBER", "JANUARY", "FEBRUARY", "MARCH",
]
MONTHS_ABBR_UPPER = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]
MONTHS_MIXED = [
    "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Jan", "Feb", "Mar", "June", "July",
]

APR_TO_DEC = {
    "full": ["APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"],
    "abbr": ["APR", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"],
    "mixed": ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "June", "July"],
}
JAN_TO_MAR = {
    "full": ["JANUARY", "FEBRUARY", "MARCH"],
    "abbr": ["JAN", "FEB", "MAR"],
    "mixed": ["Jan", "Feb", "Mar"],
}


def build_rename_map(old_fy_start: int) -> dict[str, str]:
    """Build sheet rename map. old_fy_start is the April year (e.g. 2025 for FY 2025-26)."""
    new_apr = old_fy_start + 1
    old_jan = old_fy_start + 1
    new_jan = old_fy_start + 2

    renames = {}
    suffixes_monthly = ["-EXPENSE", "-COLLECTION"]

    for m in MONTHS_SHORT_ALT:
        for sfx in suffixes_monthly:
            renames[f"{m}{old_fy_start}{sfx}"] = f"{m}{new_apr}{sfx}"

    for sfx in suffixes_monthly:
        renames[f"June{old_fy_start}{sfx}"] = f"June{new_apr}{sfx}"
        renames[f"July{old_fy_start}{sfx}"] = f"July{new_apr}{sfx}"

    for extra in ["-EXPENSE-Flatwise", "-EXPENSE-Flatwise-Trans"]:
        renames[f"July{old_fy_start}{extra}"] = f"July{new_apr}{extra}"

    for m in ["Jan", "Feb", "Mar"]:
        for sfx in suffixes_monthly:
            renames[f"{m}{old_jan}{sfx}"] = f"{m}{new_jan}{sfx}"

    renames[f"INCOME-EXPENSE-CYCLES-{old_fy_start}-{old_fy_start + 1 - 2000}"] = (
        f"INCOME-EXPENSE-CYCLES-{new_apr}-{new_apr + 1 - 2000}"
    )

    return renames


def update_formula(text: str, sorted_renames: list[tuple[str, str]]) -> str:
    for old, new in sorted_renames:
        text = text.replace(f"'{old}'!", f"'{new}'!")
        text = text.replace(f"{old}!", f"{new}!")
    return text


def update_text(text: str, old_fy_start: int) -> str:
    if not isinstance(text, str):
        return text

    new_apr = old_fy_start + 1
    old_jan = old_fy_start + 1
    new_jan = old_fy_start + 2

    old_short = f"{old_fy_start}-{(old_fy_start + 1) % 100:02d}"
    new_short = f"{new_apr}-{(new_apr + 1) % 100:02d}"
    text = text.replace(f"Apr{old_fy_start} to Mar{old_jan}", f"Apr{new_apr} to Mar{new_jan}")
    text = text.replace(old_short, new_short)
    text = text.replace(f"{old_fy_start}-{old_fy_start + 1}", f"{new_apr}-{new_jan}")

    abbr_ms = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m in abbr_ms:
        text = text.replace(f"{m}'{(old_fy_start + 2) % 100:02d}", f"{m}'{(old_fy_start + 3) % 100:02d}")
    for m in abbr_ms:
        text = text.replace(f"{m}'{(old_fy_start + 1) % 100:02d}", f"{m}'{(old_fy_start + 2) % 100:02d}")
    for m in abbr_ms:
        text = text.replace(f"{m}'{old_fy_start % 100:02d}", f"{m}'{(old_fy_start + 1) % 100:02d}")

    for group in JAN_TO_MAR.values():
        for m in group:
            text = text.replace(f"{m} {old_jan}", f"{m} {new_jan}")
            text = text.replace(f"{m}{old_jan}", f"{m}{new_jan}")

    for group in APR_TO_DEC.values():
        for m in group:
            text = text.replace(f"{m} {old_fy_start}", f"{m} {new_apr}")
            text = text.replace(f"{m}{old_fy_start}", f"{m}{new_apr}")

    return text


def is_num(v) -> bool:
    return isinstance(v, (int, float))


def detect_fy_start(wb: openpyxl.Workbook) -> int:
    """Auto-detect the April year of the current FY from sheet names."""
    for name in wb.sheetnames:
        if name.startswith("Apr") and name.endswith("-EXPENSE"):
            mid = name.replace("Apr", "").replace("-EXPENSE", "")
            if mid.isdigit():
                return int(mid)
    raise ValueError("Cannot detect FY — no 'AprYYYY-EXPENSE' sheet found")


def refresh(path: str, old_fy_start: int | None = None):
    wb = openpyxl.load_workbook(path)
    print(f"Loaded {path} ({len(wb.sheetnames)} sheets)")

    if old_fy_start is None:
        old_fy_start = detect_fy_start(wb)
    print(f"Refreshing FY {old_fy_start}-{(old_fy_start + 1) % 100:02d} → {old_fy_start + 1}-{(old_fy_start + 2) % 100:02d}")

    rename_map = build_rename_map(old_fy_start)

    # Rename sheets
    renamed = 0
    for old, new in rename_map.items():
        if old in wb.sheetnames:
            wb[old].title = new
            renamed += 1
    print(f"  Renamed {renamed} sheets")

    # Sort by length desc for safe replacement
    sorted_renames = sorted(rename_map.items(), key=lambda x: len(x[0]), reverse=True)

    # Update all cells
    updated = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, ArrayFormula):
                    v.text = update_formula(v.text, sorted_renames)
                    updated += 1
                elif isinstance(v, str):
                    if v.startswith("="):
                        new_v = update_formula(v, sorted_renames)
                        if new_v != v:
                            cell.value = new_v
                            updated += 1
                    else:
                        new_v = update_text(v, old_fy_start)
                        if new_v != v:
                            cell.value = new_v
                            updated += 1
                elif isinstance(v, datetime):
                    try:
                        cell.value = v.replace(year=v.year + 1)
                    except ValueError:
                        cell.value = v.replace(year=v.year + 1, day=28)
                    updated += 1
    print(f"  Updated {updated} cells (formulas, labels, dates)")

    # --- Zero out raw data ---

    new_apr = old_fy_start + 1
    new_jan = old_fy_start + 2

    # EXPENSE sheets
    z = 0
    expense_sheets = (
        [f"{m}{new_apr}-EXPENSE" for m in MONTHS_SHORT_ALT]
        + [f"June{new_apr}-EXPENSE", f"July{new_apr}-EXPENSE"]
        + [f"{m}{new_jan}-EXPENSE" for m in ["Jan", "Feb", "Mar"]]
    )
    for sn in expense_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            for c in [3, 4]:
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z += 1
            for c in range(12, 18):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z += 1
    print(f"  Zeroed {z} cells in EXPENSE sheets")

    # COLLECTION sheets
    z2 = 0
    collection_sheets = (
        [f"{m}{new_apr}-COLLECTION" for m in MONTHS_SHORT_ALT]
        + [f"June{new_apr}-COLLECTION", f"July{new_apr}-COLLECTION"]
        + [f"{m}{new_jan}-COLLECTION" for m in ["Jan", "Feb", "Mar"]]
    )
    for sn in collection_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            for c in [3, 5, 7, 9]:
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = None
                    z2 += 1
            for c in [4, 6, 8, 10]:
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (datetime, int, float)):
                    cell.value = None
                    z2 += 1
    print(f"  Cleared {z2} cells in COLLECTION sheets")

    # ANNUAL-EXPENSE-DETAILS
    z3 = 0
    if "ANNUAL-EXPENSE-DETAILS" in wb.sheetnames:
        ws = wb["ANNUAL-EXPENSE-DETAILS"]
        for r in range(3, 15):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z3 += 1
    print(f"  Zeroed {z3} cells in ANNUAL-EXPENSE-DETAILS")

    # INCOME-EXPENSE-CYCLES
    z4 = 0
    if "INCOME-EXPENSE-CYCLES" in wb.sheetnames:
        ws = wb["INCOME-EXPENSE-CYCLES"]
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=4)
            if is_num(cell.value):
                cell.value = 0
                z4 += 1
            for c in range(5, 53):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z4 += 1
    print(f"  Zeroed {z4} cells in INCOME-EXPENSE-CYCLES")

    # Variant INCOME-EXPENSE-CYCLES sheets
    z5 = 0
    fy_tag = f"{new_apr}-{(new_apr + 1) % 100:02d}"
    for sn in [f"INCOME-EXPENSE-CYCLES-{fy_tag}", "Copy of INCOME-EXPENSE-CYCLES 1", "Copy of INCOME-EXPENSE-CYCLES"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            for c in range(4, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z5 += 1
    print(f"  Zeroed {z5} cells in INCOME-EXPENSE-CYCLES variants")

    # July Flatwise sheets
    z6 = 0
    for sn in [f"July{new_apr}-EXPENSE-Flatwise", f"July{new_apr}-EXPENSE-Flatwise-Trans"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z6 += 1
    print(f"  Zeroed {z6} cells in July Flatwise sheets")

    # Utility sheets
    z7 = 0
    for sn, start_col in [("Sheet30", 3), ("Sheet4", 3), ("Sheet27", 2)]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z7 += 1
    if "2013-till-date-connected" in wb.sheetnames:
        ws = wb["2013-till-date-connected"]
        for r in range(4, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_num(cell.value):
                    cell.value = 0
                    z7 += 1
    print(f"  Zeroed {z7} cells in utility/archive sheets")

    wb.save(path)
    print(f"\nSaved: {path}")


def main():
    parser = argparse.ArgumentParser(description="Refresh INCOME-EXPENDITURE workbook for the next financial year")
    parser.add_argument("workbook", help="Path to the .xlsx workbook")
    parser.add_argument("--fy-start", type=int, default=None,
                        help="April year of the CURRENT FY (auto-detected if omitted)")
    parser.add_argument("--no-backup", action="store_true",
                        help="Skip creating a backup file")
    args = parser.parse_args()

    if not args.no_backup:
        backup = args.workbook.replace(".xlsx", "-backup.xlsx")
        shutil.copy2(args.workbook, backup)
        print(f"Backup: {backup}")

    refresh(args.workbook, args.fy_start)


if __name__ == "__main__":
    main()
