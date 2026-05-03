import os
import re
import json
import csv
import argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import pandas as pd
import openpyxl

def parse_filename(filepath):
    filename = os.path.basename(filepath)
    match = re.search(r'(?P<month>[a-zA-Z]+)-(?P<year>\d{4})', filename)
    if not match:
        raise ValueError(f"Could not parse month and year from filename: {filename}")
    month_str, year = match.group('month').lower(), int(match.group('year'))
    month_map = {
        'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
        'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
        'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
        'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
    }
    month = month_map.get(month_str) or month_map.get(month_str[:3])
    if not month: raise ValueError(f"Unknown month: {month_str}")
    return month, year

def get_financial_year(month, year):
    if month >= 4: return f"{year}-{str(year+1)[2:]}"
    else: return f"{year-1}-{str(year)[2:]}"

def load_mappings():
    PROJECT_ROOT = Path(__file__).resolve().parent.parent
    members_path, occupants_path = PROJECT_ROOT / 'db' / 'members.csv', PROJECT_ROOT / 'db' / 'occupants.csv'
    
    flat_to_names = defaultdict(set)      
    name_part_to_flats = defaultdict(set) 
    flat_to_name_parts = defaultdict(list) 
    flat_to_member, flat_to_occupant = {}, {}

    def process_csv(path, target_map=None):
        if not path.exists(): return
        with open(path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                flat, name = row['flat'].strip().lower(), row['name'].strip()
                if not name: continue
                if target_map is not None: target_map[flat] = name
                sub_names = [n.strip() for n in re.split(r'/|&', name)]
                for sn in sub_names:
                    if not sn: continue
                    norm_full = re.sub(r'[^a-z0-9]', '', sn.lower())
                    if norm_full: flat_to_names[flat].add(norm_full)
                    parts = [re.sub(r'[^a-z0-9]', '', p.lower()) for p in sn.split()]
                    parts = [p for p in parts if p]
                    if parts:
                        flat_to_name_parts[flat].append(parts)
                        for p in parts:
                            if len(p) >= 3: name_part_to_flats[p].add(flat)

    process_csv(members_path, flat_to_member)
    process_csv(occupants_path, flat_to_occupant)
    return flat_to_names, name_part_to_flats, flat_to_name_parts, flat_to_member, flat_to_occupant

def parse_bank_statement(filepath):
    header_row = -1
    with open(filepath, 'r') as f:
        for i, line in enumerate(f):
            if 'Txn Date' in line and 'Description' in line and 'Credit' in line:
                header_row = i
                break
    if header_row == -1: header_row = 18
    df = pd.read_csv(filepath, sep='\t', skiprows=header_row)
    df.columns = [c.strip() for c in df.columns]
    return df

def find_flat_in_description(description, flat_to_names, name_part_to_flats, flat_to_name_parts, known_flats):
    description_lower = description.lower()
    sorted_flats = sorted(known_flats, key=len, reverse=True)
    
    # 1. Flat Matching
    for f in sorted_flats:
        if re.search(r'(?i)(?:^|[^a-z0-9]|flat)' + re.escape(f) + r'(?:[^a-z0-9]|$)', description_lower): return f
        if len(f) >= 2 and f[0].isalpha() and f[1:].isdigit():
            if re.search(r'(?i)(?:^|[^a-z0-9]|flat)' + re.escape(f) + r'(?:[^0-9]|$)', description_lower): return f

    # 2. Name Matching
    norm_desc = re.sub(r'[^a-z0-9]', '', description_lower)
    desc_words = {w for w in re.split(r'[^a-z0-9]', description_lower) if w}

    # A. Full name match
    all_full_names = []
    for flat, names in flat_to_names.items():
        for n in names: all_full_names.append((n, flat))
    all_full_names.sort(key=lambda x: len(x[0]), reverse=True)
    for fn, f in all_full_names:
        if len(fn) >= 6 and fn in norm_desc: return f

    # B. Part-based match with Conflict Resolution
    potential_flats = defaultdict(float) 
    for word in desc_words:
        if len(word) < 4: continue
        if word in name_part_to_flats:
            for f in name_part_to_flats[word]: potential_flats[f] = max(potential_flats[f], 1.0)
        for part, flats in name_part_to_flats.items():
            # Relaxed match: handle "vivekdhi" matching "vivek"
            if (len(word) >= 4 and part.startswith(word)) or (len(part) >= 4 and word.startswith(part)):
                for f in flats: potential_flats[f] = max(potential_flats[f], 0.8)

    if potential_flats:
        matched = sorted(potential_flats.items(), key=lambda x: x[1], reverse=True)
        top_score = matched[0][1]
        candidates = [f for f, s in matched if s == top_score]
        if len(candidates) == 1: return candidates[0]

        # Resolve Conflicts
        best_matches = []
        max_score = -1
        for f in candidates:
            for parts in flat_to_name_parts[f]:
                score = sum(1 for p in parts if any(p==dw or (len(dw)>=3 and p.startswith(dw)) or (len(p)>=4 and dw.startswith(p)) for dw in desc_words))
                if score > max_score:
                    max_score, best_matches = score, [(f, parts, score)]
                elif score == max_score:
                    best_matches.append((f, parts, score))
        
        if len(best_matches) == 1: return best_matches[0][0]
        unique_names = {" ".join(m[1]) for m in best_matches}
        if len(unique_names) == 1: return best_matches[0][0] 
        else: return None 

    # 3. Fallback
    if '@' not in description_lower:
        for f in sorted_flats:
            if f[0].isalpha() and f in description_lower and len(f) >= 2 and f[1:].isdigit():
                return f
    return None

def update_collection_sheet(workbook_path, month, year, transactions):
    wb = openpyxl.load_workbook(workbook_path)
    sheet_name = f"{datetime(year, month, 1).strftime('%b')}{year}-COLLECTION"
    if sheet_name not in wb.sheetnames: return
    sheet = wb[sheet_name]
    flat_to_row = {str(sheet.cell(r, 1).value).strip().lower(): r for r in range(1, sheet.max_row + 1) if sheet.cell(r, 1).value}
    updates = 0
    for txn in transactions:
        flat = txn['flat']
        if flat == 'NOT MATCHED' or flat not in flat_to_row: continue
        r_idx, amt, dt = flat_to_row[flat], txn['amount'], txn['date']
        for col in range(3, 11, 2):
            a_cell, d_cell = sheet.cell(r_idx, col), sheet.cell(r_idx, col + 1)
            if a_cell.value is None:
                a_cell.value, d_cell.value, updates = amt, dt, updates + 1
                break
            elif float(str(a_cell.value).replace(',', '')) == amt and str(d_cell.value).strip() == str(dt).strip():
                break
    if updates > 0:
        try: wb.save(workbook_path)
        except PermissionError: return False
    return True

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('statement_file')
    args = parser.parse_args()
    if not os.path.exists(args.statement_file): return
    month, year = parse_filename(args.statement_file)
    fy = get_financial_year(month, year)
    PROJECT_ROOT = Path(__file__).resolve().parent.parent
    WORKBOOKS_JSON = PROJECT_ROOT / 'db' / 'workbooks.json'
    with open(WORKBOOKS_JSON, 'r') as f: workbooks = json.load(f)
    if fy not in workbooks: return
    workbook_path = PROJECT_ROOT / workbooks[fy]['workbook']
    f_names, p_flats, f_parts, f_member, f_occupant = load_mappings()
    known_flats = set(f_names.keys())
    df = parse_bank_statement(args.statement_file)
    matched, unmatched, total_credits, report_data = [], [], 0, []

    for _, row in df.iterrows():
        c_val = row.get('Credit')
        if pd.isna(c_val) or not c_val: continue
        try: amount = float(str(c_val).replace(',', ''))
        except ValueError: continue
        if amount <= 0: continue
        total_credits += 1
        desc, tx_dt = str(row.get('Description', '')), str(row.get('Txn Date', ''))
        flat = find_flat_in_description(desc, f_names, p_flats, f_parts, known_flats)
        txn = {
            'date': tx_dt, 'amount': amount, 'flat': flat or 'NOT MATCHED', 
            'member_name': f_member.get(flat, '') if flat else '', 
            'occupant_name': f_occupant.get(flat, '') if flat else '',
            'description': desc
        }
        report_data.append(txn)
        if flat: matched.append(txn)
        else: unmatched.append(txn)

    if matched: update_collection_sheet(workbook_path, month, year, matched)
    report_path = args.statement_file.replace('.xls', '_processing_report.csv')
    pd.DataFrame(report_data).to_csv(report_path, index=False)
    print(f"\nProcessed {total_credits} credits. Matched: {len(matched)}, Unmatched: {len(unmatched)}")
    if unmatched:
        print("\nUNMATCHED TRANSACTIONS:")
        for ut in unmatched[:10]:
            print(f"- {ut['date']}: {ut['amount']} | {ut['description'][:60]}...")

if __name__ == "__main__": main()
